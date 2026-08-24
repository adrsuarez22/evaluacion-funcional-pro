import streamlit as st
import pandas as pd
import altair as alt
from datetime import datetime, date
from supabase import create_client, Client
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =========================================================
# PREPARACION DATASET ESTADISTICO
# =========================================================

def normalizar_sexo(valor):
    if pd.isna(valor):
        return None
    v = str(valor).strip().lower()
    if v in ["hombre", "masculino", "m", "male"]:
        return "hombre"
    if v in ["mujer", "femenino", "f", "female"]:
        return "mujer"
    return str(valor).strip()


def preparar_df_estadistico(df):
    df = df.copy()

    columnas_fecha = ["FechaNacimiento", "Fecha"]
    for col in columnas_fecha:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.date

    if "Sexo" in df.columns:
        df["Sexo"] = df["Sexo"].apply(normalizar_sexo)

    if "FechaNacimiento" in df.columns and "Fecha" in df.columns:
        fecha_nac = pd.to_datetime(df["FechaNacimiento"], errors="coerce")
        fecha_med = pd.to_datetime(df["Fecha"], errors="coerce")
        df["Edad"] = ((fecha_med - fecha_nac).dt.days / 365.25).round(1)

    columnas_numericas = [
        "Peso_kg",
        "IMC",
        "Cintura_cm",
        "Cadera_cm",
        "ICC",
        "ICA",
        "Riesgo_ICC",
        "Riesgo_ICA",
        "Clasificacion_Abdominal",
        "Grasa_pct",
        "Musculo_kg",
        "Agua_pct",
        "Grasa_Visceral",
        "Prension",
        "Sit_to_Stand",
        "Marcha_4m",
        "SPPB",
        "TUG",
        "VO2max"
    ]

    for col in columnas_numericas:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


def preparar_dataset_longitudinal(df_estadistico):

    columnas_base = [
        "Paciente",
        "PacienteID_Ficha",
        "Sexo",
        "FechaNacimiento",
        "Edad",
        "Fecha"
    ]

    columnas_base = [c for c in columnas_base if c in df_estadistico.columns]

    value_vars = [c for c in df_estadistico.columns if c not in columnas_base]

    df_long = df_estadistico.melt(
        id_vars=columnas_base,
        value_vars=value_vars,
        var_name="Variable",
        value_name="Valor"
    )

    df_long["Valor"] = pd.to_numeric(df_long["Valor"], errors="coerce")

    df_long = df_long.dropna(subset=["Valor"])

    mapa_unidades = {
        "Peso_kg": "kg",
        "IMC": "kg/m2",
        "Cintura_cm": "cm",
        "Cadera_cm": "cm",
        "ICC": "ratio",
        "ICA": "ratio",
        "Grasa_pct": "%",
        "Musculo_kg": "kg",
        "Agua_pct": "%",
        "Grasa_Visceral": "nivel",
        "Prension": "kg"
    }

    df_long["Unidad"] = df_long["Variable"].map(mapa_unidades).fillna("")

    df_long = df_long.sort_values(["Paciente", "Fecha", "Variable"])

    return df_long


# =========================================================
# CONFIG
# =========================================================
st.set_page_config(
    page_title="Método Dra. Petratti",
    page_icon="💪",
    layout="wide"
)

st.markdown("""
<style>
[data-testid="stAppViewContainer"] {
    background-color: #f7f8fa;
}

[data-testid="stHeader"] {
    background: rgba(0, 0, 0, 0);
}

[data-testid="stToolbar"] {
    right: 1rem;
}

.block-container {
    padding-top: 1.5rem;
    padding-bottom: 2rem;
    max-width: 1400px;
}

.result-card {
    padding: 14px 16px;
    border-radius: 10px;
    font-size: 18px;
    font-weight: 700;
    margin-top: 8px;
    margin-bottom: 10px;
}

.motivo-box {
    background-color: #ffffff;
    border: 1px solid #e6e9ef;
    border-radius: 10px;
    padding: 12px 14px;
    margin-bottom: 12px;
}

.reco-box {
    background-color: #eef7ef;
    border: 1px solid #d4ead7;
    border-radius: 10px;
    padding: 12px 14px;
    margin-bottom: 14px;
}
</style>
""", unsafe_allow_html=True)

# =========================================================
# SUPABASE
# =========================================================
SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# =========================================================
# SESSION STATE / ESTADO UI
# =========================================================
if "mostrar_form_nuevo_paciente" not in st.session_state:
    st.session_state["mostrar_form_nuevo_paciente"] = False

if "paciente_id_seleccionado" not in st.session_state:
    st.session_state["paciente_id_seleccionado"] = None

if "paciente_cargado_id" not in st.session_state:
    st.session_state["paciente_cargado_id"] = None

if "busqueda_paciente" not in st.session_state:
    st.session_state["busqueda_paciente"] = ""

if "limpiar_busqueda_pendiente" not in st.session_state:
    st.session_state["limpiar_busqueda_pendiente"] = False

if "paciente_nombre_pendiente" not in st.session_state:
    st.session_state["paciente_nombre_pendiente"] = None

if "selector_paciente" not in st.session_state:
    st.session_state["selector_paciente"] = None

if "mostrar_form_editar_paciente" not in st.session_state:
    st.session_state["mostrar_form_editar_paciente"] = False


def calcular_edad_desde_fecha(fecha_nacimiento):
    if not fecha_nacimiento:
        return 0

    try:
        fecha_dt = pd.to_datetime(fecha_nacimiento, errors="coerce")
        if pd.isna(fecha_dt):
            return 0

        fecha_nac = fecha_dt.date()
        hoy = date.today()

        return (
            hoy.year
            - fecha_nac.year
            - ((hoy.month, hoy.day) < (fecha_nac.month, fecha_nac.day))
        )
    except Exception:
        return 0


def obtener_ultimo_peso_historial(df_peso):
    if df_peso is None or df_peso.empty:
        return 0.0

    df_tmp = df_peso.copy()

    if "fecha" in df_tmp.columns:
        df_tmp["fecha"] = pd.to_datetime(df_tmp["fecha"], errors="coerce")

    if "created_at" in df_tmp.columns:
        df_tmp["created_at"] = pd.to_datetime(df_tmp["created_at"], errors="coerce")

    columnas_sort = []
    ascending_sort = []

    if "fecha" in df_tmp.columns:
        columnas_sort.append("fecha")
        ascending_sort.append(False)

    if "created_at" in df_tmp.columns:
        columnas_sort.append("created_at")
        ascending_sort.append(False)

    if "id" in df_tmp.columns:
        columnas_sort.append("id")
        ascending_sort.append(False)

    if columnas_sort:
        df_tmp = df_tmp.sort_values(columnas_sort, ascending=ascending_sort)

    if "peso_kg" not in df_tmp.columns or df_tmp.empty:
        return 0.0

    peso = pd.to_numeric(df_tmp.iloc[0].get("peso_kg"), errors="coerce")

    if pd.isna(peso):
        return 0.0

    return float(peso)


def resetear_form_nuevo_paciente():
    return


def resetear_pruebas_funcionales():
    st.session_state["selector_prueba"] = "Caminata 6 minutos"
    st.session_state["valor_caminata"] = 0.0
    st.session_state["valor_prension"] = 0.0
    st.session_state["valor_silla"] = 0.0


def cargar_datos_paciente_en_widgets(paciente_actual, df_peso):
    if paciente_actual is None:
        return

    paciente_id = paciente_actual.get("id")
    peso_actual = obtener_ultimo_peso_historial(df_peso)

    resetear_pruebas_funcionales()

    st.session_state[f"peso_kg_{paciente_id}"] = float(peso_actual)
    st.session_state[f"inbody_peso_{paciente_id}"] = float(peso_actual)

    ultimo_peso = obtener_ultimo_registro(df_peso, "fecha")
    if ultimo_peso is not None:
        st.session_state[f"cintura_cm_{paciente_id}"] = float(ultimo_peso.get("cintura_cm")) if pd.notna(ultimo_peso.get("cintura_cm")) else 0.0
        st.session_state[f"cadera_cm_{paciente_id}"] = float(ultimo_peso.get("cadera_cm")) if pd.notna(ultimo_peso.get("cadera_cm")) else 0.0

    st.session_state["paciente_cargado_id"] = paciente_id
# =========================================================
# LECTURAS CACHEADAS
# =========================================================
@st.cache_data(ttl=300)
def obtener_pacientes():
    try:
        resp = (
            supabase
            .table("pacientes")
            .select("id,nombre,sexo,fecha_nacimiento,talla_m")
            .order("nombre")
            .execute()
        )
        return resp.data if resp.data else []
    except Exception as e:
        st.error(f"Error al leer pacientes: {e}")
        return []


def obtener_paciente_por_nombre(nombre):
    try:
        nombre_limpio = str(nombre).strip()

        if not nombre_limpio:
            return None

        resp = (
            supabase
            .table("pacientes")
            .select("id,nombre,sexo,fecha_nacimiento,talla_m")
            .ilike("nombre", nombre_limpio)
            .order("id", desc=True)
            .limit(1)
            .execute()
        )

        if resp.data:
            return resp.data[0]

        return None
    except Exception:
        return None


@st.cache_data(ttl=300)
def obtener_evaluaciones(paciente_id):
    try:
        resp = (
            supabase
            .table("evaluaciones")
            .select("*")
            .eq("paciente_id", paciente_id)
            .order("fecha")
            .execute()
        )
        return resp.data if resp.data else []
    except Exception:
        return []


@st.cache_data(ttl=300)
def obtener_historial_paciente(paciente_id):
    try:
        respuesta = (
            supabase.table("evaluaciones")
            .select("*")
            .eq("paciente_id", int(paciente_id))
            .order("fecha")
            .execute()
        )
        if respuesta.data:
            return pd.DataFrame(respuesta.data)
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Error al leer historial: {e}")
        return pd.DataFrame()


@st.cache_data(ttl=300)
def obtener_historial_peso(paciente_id):
    try:
        respuesta = (
            supabase.table("seguimiento_peso")
            .select("id,paciente_id,fecha,peso_kg,imc,cintura_cm,cadera_cm,icc,ica,created_at")
            .eq("paciente_id", int(paciente_id))
            .order("fecha")
            .execute()
        )

        if respuesta.data:
            return pd.DataFrame(respuesta.data)

        return pd.DataFrame(columns=["fecha", "peso_kg", "imc", "cintura_cm", "cadera_cm", "icc", "ica"])
    except Exception as e:
        st.error(f"Error al leer historial de peso: {e}")
        return pd.DataFrame(columns=["fecha", "peso_kg", "imc", "cintura_cm", "cadera_cm", "icc", "ica"])


@st.cache_data(ttl=300)
def obtener_historial_medicacion(paciente_id):
    try:
        respuesta = (
            supabase.table("medicacion_evolucion")
            .select("*")
            .eq("paciente_id", int(paciente_id))
            .order("fecha_cambio", desc=True)
            .execute()
        )

        if respuesta.data:
            return pd.DataFrame(respuesta.data)

        return pd.DataFrame(columns=[
            "fecha_cambio",
            "droga",
            "dosis",
            "unidad",
            "frecuencia",
            "via_administracion",
            "estado",
            "observaciones"
        ])
    except Exception as e:
        st.error(f"Error al leer historial de medicación: {e}")
        return pd.DataFrame(columns=[
            "fecha_cambio",
            "droga",
            "dosis",
            "unidad",
            "frecuencia",
            "via_administracion",
            "estado",
            "observaciones"
        ])


@st.cache_data(ttl=300)
def obtener_historial_inbody(paciente_id):
    try:
        respuesta = (
            supabase.table("inbody_registros")
            .select("*")
            .eq("paciente_id", int(paciente_id))
            .order("fecha", desc=True)
            .execute()
        )

        if respuesta.data:
            return pd.DataFrame(respuesta.data)

        return pd.DataFrame()

    except Exception as e:
        st.error(f"Error al leer historial de composición corporal: {e}")
        return pd.DataFrame()


# =========================================================
# TABLAS NORMATIVAS
# =========================================================
TABLA_CAMINATA_6M = {
    150: {
        40: {2.5: 436, 10: 470, 25: 511, 50: 555, 75: 592, 90: 631, 97.5: 679},
        50: {2.5: 434, 10: 468, 25: 509, 50: 553, 75: 590, 90: 629, 97.5: 677},
        60: {2.5: 414, 10: 448, 25: 489, 50: 533, 75: 570, 90: 609, 97.5: 656},
        70: {2.5: 364, 10: 397, 25: 439, 50: 483, 75: 520, 90: 558, 97.5: 606},
        80: {2.5: 313, 10: 347, 25: 388, 50: 432, 75: 469, 90: 508, 97.5: 556},
    },
    160: {
        40: {2.5: 455, 10: 489, 25: 530, 50: 574, 75: 611, 90: 650, 97.5: 697},
        50: {2.5: 453, 10: 487, 25: 528, 50: 572, 75: 609, 90: 648, 97.5: 695},
        60: {2.5: 433, 10: 466, 25: 508, 50: 552, 75: 588, 90: 627, 97.5: 675},
        70: {2.5: 382, 10: 416, 25: 457, 50: 501, 75: 538, 90: 577, 97.5: 625},
        80: {2.5: 332, 10: 366, 25: 407, 50: 451, 75: 488, 90: 526, 97.5: 574},
    },
    170: {
        40: {2.5: 474, 10: 507, 25: 549, 50: 593, 75: 629, 90: 668, 97.5: 716},
        50: {2.5: 472, 10: 505, 25: 546, 50: 590, 75: 627, 90: 666, 97.5: 714},
        60: {2.5: 451, 10: 485, 25: 526, 50: 570, 75: 607, 90: 646, 97.5: 694},
        70: {2.5: 401, 10: 435, 25: 476, 50: 520, 75: 557, 90: 595, 97.5: 643},
        80: {2.5: 351, 10: 384, 25: 425, 50: 469, 75: 506, 90: 545, 97.5: 593},
    },
    180: {
        40: {2.5: 492, 10: 526, 25: 567, 50: 611, 75: 648, 90: 687, 97.5: 735},
        50: {2.5: 490, 10: 524, 25: 565, 50: 609, 75: 646, 90: 685, 97.5: 733},
        60: {2.5: 470, 10: 503, 25: 545, 50: 589, 75: 626, 90: 664, 97.5: 712},
        70: {2.5: 419, 10: 453, 25: 494, 50: 538, 75: 575, 90: 614, 97.5: 662},
        80: {2.5: 369, 10: 403, 25: 444, 50: 488, 75: 525, 90: 564, 97.5: 611},
    },
    190: {
        40: {2.5: 511, 10: 544, 25: 586, 50: 630, 75: 667, 90: 705, 97.5: 753},
        50: {2.5: 509, 10: 542, 25: 584, 50: 628, 75: 665, 90: 703, 97.5: 751},
        60: {2.5: 488, 10: 522, 25: 563, 50: 607, 75: 644, 90: 683, 97.5: 731},
        70: {2.5: 438, 10: 472, 25: 513, 50: 557, 75: 594, 90: 633, 97.5: 680},
        80: {2.5: 388, 10: 421, 25: 463, 50: 507, 75: 544, 90: 582, 97.5: 630},
    }
}

TABLA_PRENSION = {
    "Hombre": {
        "20-24": {5: 33.9, 10: 36.8, 20: 40.5, 30: 43.2, 40: 45.7, 50: 48.0, 60: 50.4, 70: 52.9, 80: 56.0, 90: 60.1, 95: 63.6},
        "25-29": {5: 35.5, 10: 38.5, 20: 42.1, 30: 44.8, 40: 47.1, 50: 49.3, 60: 51.5, 70: 53.9, 80: 56.7, 90: 60.7, 95: 64.0},
        "30-34": {5: 35.0, 10: 38.3, 20: 42.2, 30: 45.0, 40: 47.4, 50: 49.7, 60: 52.0, 70: 54.4, 80: 57.4, 90: 61.5, 95: 64.9},
        "35-39": {5: 33.8, 10: 37.3, 20: 41.5, 30: 44.5, 40: 47.1, 50: 49.5, 60: 51.9, 70: 54.4, 80: 57.5, 90: 61.8, 95: 65.3},
        "40-44": {5: 32.3, 10: 36.0, 20: 40.4, 30: 43.6, 40: 46.3, 50: 48.8, 60: 51.2, 70: 53.9, 80: 57.1, 90: 61.5, 95: 65.1},
        "45-49": {5: 30.6, 10: 34.4, 20: 39.0, 30: 42.3, 40: 45.1, 50: 47.6, 60: 50.2, 70: 52.9, 80: 56.2, 90: 60.7, 95: 64.4},
        "50-54": {5: 28.9, 10: 32.8, 20: 37.4, 30: 40.7, 40: 43.5, 50: 46.2, 60: 48.8, 70: 51.6, 80: 54.8, 90: 59.4, 95: 63.1},
        "55-59": {5: 27.2, 10: 31.0, 20: 35.6, 30: 38.9, 40: 41.7, 50: 44.4, 60: 47.0, 70: 49.8, 80: 53.1, 90: 57.7, 95: 61.4},
        "60-64": {5: 25.5, 10: 29.1, 20: 33.6, 30: 36.9, 40: 39.7, 50: 42.4, 60: 45.0, 70: 47.8, 80: 51.1, 90: 55.6, 95: 59.3},
        "65-69": {5: 23.7, 10: 27.2, 20: 31.5, 30: 34.7, 40: 37.5, 50: 40.1, 60: 42.8, 70: 45.6, 80: 48.8, 90: 53.2, 95: 56.8},
        "70-74": {5: 21.9, 10: 25.2, 20: 29.3, 30: 32.4, 40: 35.1, 50: 37.7, 60: 40.3, 70: 43.1, 80: 46.3, 90: 50.6, 95: 54.1},
        "75-79": {5: 20.0, 10: 23.1, 20: 27.0, 30: 29.9, 40: 32.5, 50: 35.1, 60: 37.6, 70: 40.3, 80: 43.5, 90: 47.7, 95: 51.1},
        "80-84": {5: 18.0, 10: 20.8, 20: 24.5, 30: 27.3, 40: 29.8, 50: 32.3, 60: 34.8, 70: 37.5, 80: 40.5, 90: 44.7, 95: 48.0},
        "85-89": {5: 15.9, 10: 18.5, 20: 21.9, 30: 24.6, 40: 27.0, 50: 29.4, 60: 31.8, 70: 34.4, 80: 37.4, 90: 41.5, 95: 44.6},
        "90-94": {5: 13.7, 10: 16.1, 20: 19.2, 30: 21.7, 40: 24.0, 50: 26.3, 60: 28.7, 70: 31.2, 80: 34.2, 90: 38.1, 95: 41.2},
        "95-99": {5: 11.3, 10: 13.5, 20: 16.4, 30: 18.8, 40: 20.9, 50: 23.1, 60: 25.4, 70: 27.9, 80: 30.8, 90: 34.6, 95: 37.5},
        "+100": {5: 8.8, 10: 10.8, 20: 13.5, 30: 15.7, 40: 17.8, 50: 19.8, 60: 22.0, 70: 24.5, 80: 27.2, 90: 30.9, 95: 33.8},
    },
    "Mujer": {
        "20-24": {5: 19.7, 10: 21.7, 20: 24.0, 30: 25.7, 40: 27.2, 50: 28.6, 60: 30.0, 70: 31.6, 80: 33.6, 90: 36.6, 95: 39.1},
        "25-29": {5: 20.0, 10: 22.0, 20: 24.5, 30: 26.3, 40: 27.9, 50: 29.4, 60: 30.9, 70: 32.6, 80: 34.6, 90: 37.4, 95: 39.7},
        "30-34": {5: 19.6, 10: 21.8, 20: 24.4, 30: 26.4, 40: 28.1, 50: 29.7, 60: 31.3, 70: 33.1, 80: 35.2, 90: 38.0, 95: 40.4},
        "35-39": {5: 19.0, 10: 21.3, 20: 24.1, 30: 26.2, 40: 28.0, 50: 29.7, 60: 31.4, 70: 33.2, 80: 35.4, 90: 38.4, 95: 40.8},
        "40-44": {5: 18.3, 10: 20.7, 20: 23.7, 30: 25.8, 40: 27.6, 50: 29.4, 60: 31.1, 70: 33.0, 80: 35.2, 90: 38.3, 95: 40.8},
        "45-49": {5: 17.6, 10: 20.1, 20: 23.1, 30: 25.2, 40: 27.1, 50: 28.9, 60: 30.6, 70: 32.5, 80: 34.8, 90: 37.9, 95: 40.4},
        "50-54": {5: 16.9, 10: 19.4, 20: 22.4, 30: 24.5, 40: 26.4, 50: 28.2, 60: 29.9, 70: 31.8, 80: 34.0, 90: 37.1, 95: 39.7},
        "55-59": {5: 16.1, 10: 18.5, 20: 21.5, 30: 23.7, 40: 25.5, 50: 27.3, 60: 29.0, 70: 30.9, 80: 33.0, 90: 36.1, 95: 38.6},
        "60-64": {5: 15.2, 10: 17.6, 20: 20.6, 30: 22.7, 40: 24.5, 50: 26.2, 60: 27.9, 70: 29.7, 80: 31.8, 90: 34.9, 95: 37.4},
        "65-69": {5: 14.3, 10: 16.6, 20: 19.5, 30: 21.6, 40: 23.3, 50: 25.0, 60: 26.6, 70: 28.4, 80: 30.5, 90: 33.4, 95: 35.8},
        "70-74": {5: 13.2, 10: 15.5, 20: 18.3, 30: 20.3, 40: 22.0, 50: 23.6, 60: 25.2, 70: 26.9, 80: 28.9, 90: 31.8, 95: 34.1},
        "75-79": {5: 12.0, 10: 14.3, 20: 17.0, 30: 18.9, 40: 20.5, 50: 22.1, 60: 23.6, 70: 25.2, 80: 27.2, 90: 29.9, 95: 32.2},
        "80-84": {5: 10.7, 10: 12.9, 20: 15.5, 30: 17.4, 40: 18.9, 50: 20.4, 60: 21.9, 70: 23.5, 80: 25.3, 90: 28.0, 95: 30.2},
        "85-89": {5: 9.3, 10: 11.4, 20: 13.9, 30: 15.7, 40: 17.2, 50: 18.6, 60: 20.0, 70: 21.5, 80: 23.3, 90: 25.9, 95: 28.0},
        "90-94": {5: 7.8, 10: 9.8, 20: 12.2, 30: 13.9, 40: 15.3, 50: 16.7, 60: 18.0, 70: 19.5, 80: 21.2, 90: 23.6, 95: 25.7},
        "95-99": {5: 6.1, 10: 8.0, 20: 10.3, 30: 11.9, 40: 13.3, 50: 14.6, 60: 15.9, 70: 17.3, 80: 18.9, 90: 21.2, 95: 23.2},
        "+100": {5: 4.2, 10: 6.1, 20: 8.3, 30: 9.8, 40: 11.2, 50: 12.4, 60: 13.6, 70: 14.9, 80: 16.5, 90: 18.7, 95: 20.6},
    }
}

TABLA_SILLA = {
    "Hombre": {
        "65-69": {10: 12, 20: 13, 30: 14, 40: 15, 50: 16, 60: 16, 70: 17, 80: 19, 90: 21, 100: 28},
        "70-74": {10: 11, 20: 13, 30: 14, 40: 15, 50: 15, 60: 16, 70: 17, 80: 18, 90: 20, 100: 29},
        "75-79": {10: 10, 20: 12, 30: 13, 40: 14, 50: 14, 60: 15, 70: 16, 80: 17, 90: 19, 100: 25},
        "80-84": {10: 9, 20: 10, 30: 11, 40: 12, 50: 14, 60: 15, 70: 16, 80: 17, 90: 18, 100: 22},
        "+84": {10: 9, 20: 9, 30: 12, 40: 13, 50: 14, 60: 14, 70: 16, 80: 18, 90: 20, 100: 21},
    },
    "Mujer": {
        "65-69": {10: 11, 20: 12, 30: 13, 40: 14, 50: 15, 60: 15, 70: 16, 80: 17, 90: 19, 100: 30},
        "70-74": {10: 10, 20: 12, 30: 12, 40: 13, 50: 14, 60: 15, 70: 16, 80: 17, 90: 19, 100: 27},
        "75-79": {10: 10, 20: 11, 30: 12, 40: 13, 50: 14, 60: 14, 70: 15, 80: 16, 90: 18, 100: 24},
        "80-84": {10: 9, 20: 10, 30: 11, 40: 12, 50: 13, 60: 14, 70: 15, 80: 16, 90: 18, 100: 24},
        "+84": {10: 6, 20: 8, 30: 9, 40: 11, 50: 12, 60: 14, 70: 14, 80: 16, 90: 17, 100: 18},
    }
}

# =========================================================
# BASE DE DATOS - ESCRITURAS
# =========================================================
def limpiar_cache():
    st.cache_data.clear()


def guardar_evaluacion(paciente_id, paciente_nombre, sexo, edad, prueba, valor_medido, percentil, clasificacion):
    if paciente_id is None:
        raise ValueError("No se encontró el id del paciente.")

    payload = {
        "fecha": datetime.now().strftime("%Y-%m-%d"),
        "paciente": str(paciente_nombre).strip(),
        "sexo": str(sexo).strip().lower(),
        "edad": int(edad),
        "prueba": str(prueba).strip(),
        "valor_medido": float(valor_medido),
        "percentil": round(float(percentil), 1) if percentil is not None else None,
        "clasificacion": str(clasificacion).strip(),
        "paciente_id": int(paciente_id)
    }
    resp = supabase.table("evaluaciones").insert(payload).execute()
    limpiar_cache()
    return resp


def guardar_paciente(nombre, sexo, fecha_nacimiento, talla_m):
    nombre_limpio = str(nombre).strip()
    sexo_limpio = str(sexo).strip().lower()

    if not nombre_limpio:
        raise ValueError("El nombre del paciente está vacío.")

    if talla_m is None or float(talla_m) <= 0:
        raise ValueError("La talla debe ser mayor a 0.")

    respuesta = supabase.table("pacientes").select("id,nombre").execute()
    existentes = respuesta.data if respuesta.data else []

    for p in existentes:
        if str(p["nombre"]).strip().lower() == nombre_limpio.lower():
            raise ValueError("Ese paciente ya existe.")

    payload = {
        "nombre": nombre_limpio,
        "sexo": sexo_limpio,
        "fecha_nacimiento": str(fecha_nacimiento),
        "talla_m": round(float(talla_m), 2)
    }

    resp = supabase.table("pacientes").insert(payload).execute()
    limpiar_cache()
    return resp


def actualizar_paciente(paciente_id, nombre, sexo, fecha_nacimiento, talla_m):
    nombre_limpio = str(nombre).strip()
    sexo_limpio = str(sexo).strip().lower()

    if not nombre_limpio:
        raise ValueError("El nombre del paciente está vacío.")

    if talla_m is None or float(talla_m) <= 0:
        raise ValueError("La talla debe ser mayor a 0.")

    respuesta = supabase.table("pacientes").select("id,nombre").execute()
    existentes = respuesta.data if respuesta.data else []

    for p in existentes:
        if int(p["id"]) != int(paciente_id) and str(p["nombre"]).strip().lower() == nombre_limpio.lower():
            raise ValueError("Ya existe otro paciente con ese nombre.")

    payload = {
        "nombre": nombre_limpio,
        "sexo": sexo_limpio,
        "fecha_nacimiento": str(fecha_nacimiento),
        "talla_m": round(float(talla_m), 2)
    }

    resp = (
        supabase
        .table("pacientes")
        .update(payload)
        .eq("id", int(paciente_id))
        .execute()
    )
    limpiar_cache()
    return resp


def eliminar_paciente(paciente_id):
    resp_del = supabase.table("pacientes").delete().eq("id", paciente_id).execute()
    limpiar_cache()
    return resp_del


def guardar_peso(paciente_id, fecha_medicion, peso_kg, talla_m, cintura_cm=None, cadera_cm=None):
    if paciente_id is None:
        raise ValueError("No se encontró el id del paciente.")

    if talla_m is None or float(talla_m) <= 0:
        raise ValueError("El paciente no tiene una talla válida cargada.")

    if peso_kg is None or float(peso_kg) <= 0:
        raise ValueError("El peso debe ser mayor a 0.")

    if cintura_cm is not None and float(cintura_cm) < 0:
        raise ValueError("La cintura no puede ser negativa.")

    if cadera_cm is not None and float(cadera_cm) < 0:
        raise ValueError("La cadera no puede ser negativa.")

    imc = round(float(peso_kg) / (float(talla_m) ** 2), 2)
    icc = calcular_icc(cintura_cm, cadera_cm)
    ica = calcular_ica(cintura_cm, talla_m)
    fecha_txt = str(fecha_medicion)

    existente = (
        supabase.table("seguimiento_peso")
        .select("id")
        .eq("paciente_id", int(paciente_id))
        .eq("fecha", fecha_txt)
        .execute()
    )

    payload = {
        "paciente_id": int(paciente_id),
        "fecha": fecha_txt,
        "peso_kg": float(peso_kg),
        "imc": imc,
        "cintura_cm": float(cintura_cm) if cintura_cm is not None and float(cintura_cm) > 0 else None,
        "cadera_cm": float(cadera_cm) if cadera_cm is not None and float(cadera_cm) > 0 else None,
        "icc": float(icc) if icc is not None else None,
        "ica": float(ica) if ica is not None else None
    }

    if existente.data:
        id_existente = existente.data[0]["id"]
        resp = (
            supabase.table("seguimiento_peso")
            .update(payload)
            .eq("id", id_existente)
            .execute()
        )
    else:
        resp = supabase.table("seguimiento_peso").insert(payload).execute()

    limpiar_cache()
    return resp


def guardar_medicacion(
    paciente_id,
    fecha_cambio,
    droga,
    dosis,
    unidad,
    frecuencia,
    via_administracion,
    estado,
    observaciones
):
    if paciente_id is None:
        raise ValueError("No se encontró el id del paciente.")

    if not str(droga).strip():
        raise ValueError("La droga es obligatoria.")

    payload = {
        "paciente_id": int(paciente_id),
        "fecha_cambio": str(fecha_cambio),
        "droga": str(droga).strip(),
        "dosis": float(dosis) if dosis is not None else None,
        "unidad": str(unidad).strip() if unidad else None,
        "frecuencia": str(frecuencia).strip() if frecuencia else None,
        "via_administracion": str(via_administracion).strip() if via_administracion else None,
        "estado": str(estado).strip() if estado else "Activa",
        "observaciones": str(observaciones).strip() if observaciones else None
    }

    resp = supabase.table("medicacion_evolucion").insert(payload).execute()
    limpiar_cache()
    return resp


def guardar_inbody(
    paciente_id,
    fecha_estudio,
    peso_kg,
    talla_m,
    grasa_corporal_pct,
    masa_muscular_kg,
    agua_corporal_pct,
    grasa_visceral,
    metabolismo_basal,
    observaciones
):
    if paciente_id is None:
        raise ValueError("No se encontró el id del paciente.")

    imc_calculado = round(float(peso_kg) / (float(talla_m) ** 2), 2) if peso_kg and talla_m else None

    payload = {
        "paciente_id": int(paciente_id),
        "fecha": str(fecha_estudio),
        "peso_kg": float(peso_kg) if peso_kg is not None else None,
        "imc": float(imc_calculado) if imc_calculado is not None else None,
        "grasa_corporal_pct": float(grasa_corporal_pct) if grasa_corporal_pct is not None else None,
        "masa_muscular_kg": float(masa_muscular_kg) if masa_muscular_kg is not None else None,
        "agua_corporal_pct": float(agua_corporal_pct) if agua_corporal_pct is not None else None,
        "grasa_visceral": float(grasa_visceral) if grasa_visceral is not None else None,
        "metabolismo_basal": float(metabolismo_basal) if metabolismo_basal is not None else None,
        "observaciones": str(observaciones).strip() if observaciones else None
    }

    resp = supabase.table("inbody_registros").insert(payload).execute()
    limpiar_cache()
    return resp


def eliminar_evaluacion(id_registro):
    resp = supabase.table("evaluaciones").delete().eq("id", id_registro).execute()
    limpiar_cache()
    return resp


def eliminar_registro_peso(id_registro):
    resp = supabase.table("seguimiento_peso").delete().eq("id", id_registro).execute()
    limpiar_cache()
    return resp


def eliminar_registro_corporal(id_registro):
    resp = supabase.table("inbody_registros").delete().eq("id", id_registro).execute()
    limpiar_cache()
    return resp


def obtener_ultimo_id_peso(df_peso):
    if df_peso is None or df_peso.empty or "id" not in df_peso.columns:
        return None

    df_tmp = df_peso.copy()

    if "fecha" in df_tmp.columns:
        df_tmp["fecha"] = pd.to_datetime(df_tmp["fecha"], errors="coerce")
    if "created_at" in df_tmp.columns:
        df_tmp["created_at"] = pd.to_datetime(df_tmp["created_at"], errors="coerce")

    columnas_sort = []
    ascending_sort = []

    if "fecha" in df_tmp.columns:
        columnas_sort.append("fecha")
        ascending_sort.append(False)

    if "created_at" in df_tmp.columns:
        columnas_sort.append("created_at")
        ascending_sort.append(False)

    columnas_sort.append("id")
    ascending_sort.append(False)

    df_tmp = df_tmp.sort_values(columnas_sort, ascending=ascending_sort)
    return int(df_tmp.iloc[0]["id"])


def obtener_ultimo_id_inbody(df_inbody):
    if df_inbody is None or df_inbody.empty or "id" not in df_inbody.columns:
        return None

    df_tmp = df_inbody.copy()

    if "fecha" in df_tmp.columns:
        df_tmp["fecha"] = pd.to_datetime(df_tmp["fecha"], errors="coerce")
        df_tmp = df_tmp.sort_values(["fecha", "id"], ascending=[False, False])
    else:
        df_tmp = df_tmp.sort_values("id", ascending=False)

    return int(df_tmp.iloc[0]["id"])


# =========================================================
# UTILIDADES PACIENTE / FICHA
# =========================================================
def construir_ficha_paciente(paciente_actual, df_historial):
    if paciente_actual is None:
        return {
            "id": None,
            "nombre": "-",
            "sexo": "-",
            "fecha_nacimiento": None,
            "talla_m": None,
            "cantidad_evaluaciones": 0,
            "ultima_fecha": "-",
            "ultima_clasificacion": "-",
            "ultima_prueba": "-"
        }

    cantidad_evaluaciones = 0
    ultima_fecha = "-"
    ultima_clasificacion = "-"
    ultima_prueba = "-"

    if df_historial is not None and not df_historial.empty:
        df_tmp = df_historial.copy()

        if "fecha" in df_tmp.columns:
            df_tmp["fecha"] = pd.to_datetime(df_tmp["fecha"], errors="coerce")
            df_tmp = df_tmp.sort_values("fecha", ascending=False)

        cantidad_evaluaciones = len(df_tmp)

        if "fecha" in df_tmp.columns and pd.notna(df_tmp.iloc[0]["fecha"]):
            ultima_fecha = df_tmp.iloc[0]["fecha"].strftime("%d-%m-%Y")

        if "clasificacion" in df_tmp.columns and pd.notna(df_tmp.iloc[0]["clasificacion"]):
            ultima_clasificacion = str(df_tmp.iloc[0]["clasificacion"])

        if "prueba" in df_tmp.columns and pd.notna(df_tmp.iloc[0]["prueba"]):
            ultima_prueba = str(df_tmp.iloc[0]["prueba"])

    return {
        "id": paciente_actual.get("id"),
        "nombre": paciente_actual.get("nombre", "-"),
        "sexo": paciente_actual.get("sexo", "-"),
        "fecha_nacimiento": paciente_actual.get("fecha_nacimiento"),
        "talla_m": paciente_actual.get("talla_m"),
        "cantidad_evaluaciones": cantidad_evaluaciones,
        "ultima_fecha": ultima_fecha,
        "ultima_clasificacion": ultima_clasificacion,
        "ultima_prueba": ultima_prueba
    }


def preparar_df_exportacion(df):
    if df is None or df.empty:
        return pd.DataFrame()

    df_out = df.copy()

    for col in df_out.columns:
        if "fecha" in col.lower() or "created" in col.lower():
            try:
                df_out[col] = pd.to_datetime(df_out[col], errors="coerce")
                df_out[col] = df_out[col].dt.strftime("%Y-%m-%d")
            except Exception:
                pass

    return df_out


# =========================================================
# UTILIDADES CLÍNICAS - COMPOSICIÓN CORPORAL
# =========================================================
def clasificacion_grasa_corporal(sexo, grasa_pct):
    sexo = str(sexo).strip().lower()

    if grasa_pct is None or pd.isna(grasa_pct):
        return "Sin clasificar"

    if sexo == "hombre":
        if grasa_pct < 8:
            return "Muy bajo"
        elif grasa_pct < 19:
            return "Normal"
        elif grasa_pct <= 25:
            return "Alto"
        else:
            return "Obesidad"

    if sexo == "mujer":
        if grasa_pct < 21:
            return "Muy bajo"
        elif grasa_pct < 33:
            return "Normal"
        elif grasa_pct <= 39:
            return "Alto"
        else:
            return "Obesidad"

    return "Sin clasificar"


def clasificacion_agua_corporal(sexo, agua_pct):
    sexo = str(sexo).strip().lower()

    if agua_pct is None or pd.isna(agua_pct):
        return "Sin clasificar"

    if sexo == "hombre":
        if agua_pct < 50:
            return "Bajo"
        elif agua_pct <= 65:
            return "Normal"
        else:
            return "Alto"

    if sexo == "mujer":
        if agua_pct < 45:
            return "Bajo"
        elif agua_pct <= 60:
            return "Normal"
        else:
            return "Alto"

    return "Sin clasificar"


def clasificacion_grasa_visceral(grasa_visceral):
    if grasa_visceral is None or pd.isna(grasa_visceral):
        return "Sin clasificar"

    if grasa_visceral <= 9:
        return "Normal"
    elif grasa_visceral <= 14:
        return "Alto"
    else:
        return "Muy alto"


def calcular_masa_muscular_relativa_pct(peso_kg, masa_muscular_kg):
    if peso_kg is None or masa_muscular_kg is None:
        return None
    if pd.isna(peso_kg) or pd.isna(masa_muscular_kg):
        return None
    if float(peso_kg) <= 0:
        return None
    return round((float(masa_muscular_kg) / float(peso_kg)) * 100, 2)


def clasificacion_masa_muscular_relativa(sexo, musculo_relativo_pct):
    sexo = str(sexo).strip().lower()

    if musculo_relativo_pct is None or pd.isna(musculo_relativo_pct):
        return "Sin clasificar"

    if sexo == "hombre":
        if musculo_relativo_pct < 33:
            return "Bajo"
        elif musculo_relativo_pct <= 39:
            return "Normal"
        else:
            return "Alto"

    if sexo == "mujer":
        if musculo_relativo_pct < 24:
            return "Bajo"
        elif musculo_relativo_pct <= 30:
            return "Normal"
        else:
            return "Alto"

    return "Sin clasificar"


def color_estado_corporal(estado):
    mapa = {
        "Normal": ("#2e7d32", "#ffffff"),
        "Bajo peso": ("#1976d2", "#ffffff"),
        "Riesgo sarcopénico": ("#ef6c00", "#ffffff"),
        "Sobrepeso": ("#f9a825", "#1f1f1f"),
        "Sobrepeso muscular": ("#00897b", "#ffffff"),
        "Obesidad": ("#c62828", "#ffffff"),
        "Riesgo cardiometabólico": ("#ad1457", "#ffffff"),
        "Riesgo cardiometabólico moderado": ("#6a1b9a", "#ffffff"),
        "Sin clasificar": ("#757575", "#ffffff")
    }
    return mapa.get(estado, ("#757575", "#ffffff"))


def generar_recomendacion_corporal(estado, clasif_grasa, clasif_visceral, clasif_musculo):
    if estado in ["Obesidad", "Riesgo cardiometabólico", "Riesgo cardiometabólico moderado"]:
        return "Programa de reducción de grasa + ejercicio de fuerza + control cardiometabólico."
    if estado == "Riesgo sarcopénico":
        return "Priorizar ejercicio de fuerza, aumento de masa muscular y seguimiento funcional."
    if estado == "Sobrepeso muscular":
        return "Mantener masa muscular, controlar evolución y ajustar plan nutricional según objetivo."
    if estado == "Sobrepeso":
        return "Plan de control de peso con actividad física regular y seguimiento de composición corporal."
    if estado == "Bajo peso":
        return "Evaluar aporte nutricional y preservar o mejorar masa muscular."
    if estado == "Normal":
        if clasif_musculo == "Bajo":
            return "Estado general aceptable, pero conviene reforzar trabajo de fuerza."
        return "Mantener hábitos actuales y seguimiento periódico."
    return "Completar datos clínicos y repetir control."


def calcular_icc(cintura_cm, cadera_cm):
    if cintura_cm is None or cadera_cm is None:
        return None
    if pd.isna(cintura_cm) or pd.isna(cadera_cm):
        return None
    if float(cintura_cm) <= 0 or float(cadera_cm) <= 0:
        return None
    return round(float(cintura_cm) / float(cadera_cm), 4)


def calcular_ica(cintura_cm, talla_m):
    if cintura_cm is None or talla_m is None:
        return None
    if pd.isna(cintura_cm) or pd.isna(talla_m):
        return None
    talla_cm = float(talla_m) * 100
    if float(cintura_cm) <= 0 or talla_cm <= 0:
        return None
    return round(float(cintura_cm) / talla_cm, 4)


def clasificacion_icc(sexo, icc):
    sexo = str(sexo).strip().lower()

    if icc is None or pd.isna(icc):
        return "Sin clasificar"

    if sexo == "mujer":
        return "Riesgo aumentado" if float(icc) > 0.85 else "Riesgo no aumentado"

    if sexo == "hombre":
        return "Riesgo aumentado" if float(icc) > 0.90 else "Riesgo no aumentado"

    return "Sin clasificar"


def clasificacion_ica(ica):
    if ica is None or pd.isna(ica):
        return "Sin clasificar"

    ica = float(ica)

    if ica < 0.5:
        return "Riesgo bajo"
    elif ica < 0.6:
        return "Riesgo cardiometabólico aumentado"
    else:
        return "Riesgo alto"


def clasificacion_obesidad_abdominal(sexo, icc, ica):
    riesgo_icc = clasificacion_icc(sexo, icc)
    riesgo_ica = clasificacion_ica(ica)

    if riesgo_icc == "Sin clasificar" and riesgo_ica == "Sin clasificar":
        return "Sin clasificar"

    if riesgo_ica == "Riesgo alto":
        return "Obesidad abdominal / riesgo alto"

    if riesgo_icc == "Riesgo aumentado" and riesgo_ica == "Riesgo cardiometabólico aumentado":
        return "Obesidad abdominal / riesgo aumentado"

    if riesgo_icc == "Riesgo aumentado" or riesgo_ica == "Riesgo cardiometabólico aumentado":
        return "Riesgo abdominal aumentado"

    return "Riesgo abdominal no aumentado"


def clasificar_imc(imc_calculado):
    if imc_calculado < 18.5:
        return "Bajo peso", "🔵"
    elif imc_calculado < 25:
        return "Normal", "🟢"
    elif imc_calculado < 30:
        return "Sobrepeso", "🟡"
    else:
        return "Obesidad", "🔴"


def evaluar_perfil_morfofuncional(sexo, peso_kg, talla_m, grasa_pct, masa_muscular_kg, agua_pct, grasa_visceral):
    imc = round(float(peso_kg) / (float(talla_m) ** 2), 2) if peso_kg and talla_m else None
    clasif_imc = clasificar_imc(imc)[0] if imc is not None else "Sin clasificar"
    clasif_grasa = clasificacion_grasa_corporal(sexo, grasa_pct)
    clasif_agua = clasificacion_agua_corporal(sexo, agua_pct)
    clasif_visceral = clasificacion_grasa_visceral(grasa_visceral)
    musculo_rel_pct = calcular_masa_muscular_relativa_pct(peso_kg, masa_muscular_kg)
    clasif_musculo = clasificacion_masa_muscular_relativa(sexo, musculo_rel_pct)

    motivos = []

    if clasif_imc == "Bajo peso":
        motivos.append("IMC en rango de bajo peso")
    elif clasif_imc == "Sobrepeso":
        motivos.append("IMC en rango de sobrepeso")
    elif clasif_imc == "Obesidad":
        motivos.append("IMC en rango de obesidad")

    if clasif_grasa in ["Alto", "Obesidad"]:
        motivos.append("% grasa corporal elevada")

    if clasif_visceral in ["Alto", "Muy alto"]:
        motivos.append(f"grasa visceral {clasif_visceral.lower()}")

    if clasif_musculo == "Bajo":
        motivos.append("masa muscular relativa baja")
    elif clasif_musculo == "Normal":
        motivos.append("masa muscular relativa normal")
    elif clasif_musculo == "Alto":
        motivos.append("masa muscular relativa alta")

    estado = "Normal"

    if clasif_visceral == "Muy alto":
        estado = "Riesgo cardiometabólico"
    elif clasif_visceral == "Alto" and clasif_grasa in ["Alto", "Obesidad"]:
        estado = "Riesgo cardiometabólico moderado"
    elif clasif_imc == "Obesidad" or (clasif_imc in ["Sobrepeso", "Obesidad"] and clasif_grasa == "Obesidad"):
        estado = "Obesidad"
    elif clasif_imc == "Sobrepeso" and clasif_grasa in ["Alto", "Obesidad"] and clasif_musculo != "Alto":
        estado = "Sobrepeso"
    elif clasif_imc == "Sobrepeso" and clasif_musculo == "Alto" and clasif_grasa == "Normal":
        estado = "Sobrepeso muscular"
    elif clasif_imc == "Normal" and clasif_musculo == "Bajo":
        estado = "Riesgo sarcopénico"
    elif clasif_imc == "Bajo peso":
        estado = "Bajo peso"
    else:
        estado = "Normal"

    recomendacion = generar_recomendacion_corporal(
        estado=estado,
        clasif_grasa=clasif_grasa,
        clasif_visceral=clasif_visceral,
        clasif_musculo=clasif_musculo
    )

    return {
        "imc": imc,
        "clasif_imc": clasif_imc,
        "clasif_grasa": clasif_grasa,
        "clasif_agua": clasif_agua,
        "clasif_visceral": clasif_visceral,
        "musculo_rel_pct": musculo_rel_pct,
        "clasif_musculo": clasif_musculo,
        "estado": estado,
        "motivos": motivos,
        "recomendacion": recomendacion
    }


def enriquecer_historial_corporal(df, sexo, talla_m):
    if df is None or df.empty:
        return pd.DataFrame()

    df = df.copy()

    for col in [
        "peso_kg",
        "imc",
        "grasa_corporal_pct",
        "masa_muscular_kg",
        "agua_corporal_pct",
        "grasa_visceral",
        "metabolismo_basal"
    ]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    diagnosticos = []
    sugerencias = []
    motivos_lista = []
    clasif_imc_lista = []
    clasif_grasa_lista = []
    clasif_agua_lista = []
    clasif_visceral_lista = []
    musculo_rel_lista = []
    clasif_musculo_lista = []

    for _, row in df.iterrows():
        res = evaluar_perfil_morfofuncional(
            sexo=sexo,
            peso_kg=row.get("peso_kg"),
            talla_m=talla_m,
            grasa_pct=row.get("grasa_corporal_pct"),
            masa_muscular_kg=row.get("masa_muscular_kg"),
            agua_pct=row.get("agua_corporal_pct"),
            grasa_visceral=row.get("grasa_visceral")
        )

        diagnosticos.append(res["estado"])
        sugerencias.append(res["recomendacion"])
        motivos_lista.append(" | ".join(res["motivos"]) if res["motivos"] else "")
        clasif_imc_lista.append(res["clasif_imc"])
        clasif_grasa_lista.append(res["clasif_grasa"])
        clasif_agua_lista.append(res["clasif_agua"])
        clasif_visceral_lista.append(res["clasif_visceral"])
        musculo_rel_lista.append(res["musculo_rel_pct"])
        clasif_musculo_lista.append(res["clasif_musculo"])

    df["diagnostico_corporal"] = diagnosticos
    df["sugerencia_corporal"] = sugerencias
    df["motivos_corporal"] = motivos_lista
    df["clasif_imc"] = clasif_imc_lista
    df["clasif_grasa"] = clasif_grasa_lista
    df["clasif_agua"] = clasif_agua_lista
    df["clasif_visceral"] = clasif_visceral_lista
    df["musculo_rel_pct"] = musculo_rel_lista
    df["clasif_musculo"] = clasif_musculo_lista

    return df


def agregar_identificacion_paciente(df, ficha, origen=""):
    if df is None or df.empty:
        return pd.DataFrame(columns=[
            "Paciente",
            "PacienteID_Ficha",
            "Sexo",
            "Talla_m",
            "Origen"
        ])

    df_out = df.copy()

    rename_map = {}
    if "id" in df_out.columns:
        rename_map["id"] = "RegistroID"
    if "paciente_id" in df_out.columns:
        rename_map["paciente_id"] = "PacienteID_Registro"
    df_out = df_out.rename(columns=rename_map)

    columnas_identificacion = {
        "Paciente": ficha.get("nombre"),
        "PacienteID_Ficha": ficha.get("id"),
        "Sexo": ficha.get("sexo"),
        "Talla_m": ficha.get("talla_m"),
        "Origen": origen
    }

    for nombre_col, valor in reversed(list(columnas_identificacion.items())):
        if nombre_col in df_out.columns:
            df_out.drop(columns=[nombre_col], inplace=True)
        df_out.insert(0, nombre_col, valor)

    return df_out


def generar_df_analisis_cientifico(ficha, df_peso, df_inbody, df_eval, df_medicacion):
    filas = []

    sexo = ficha.get("sexo")
    talla_m = ficha.get("talla_m")

    def obtener_tratamiento_vigente(fecha_referencia):
        if df_medicacion is None or df_medicacion.empty or pd.isna(fecha_referencia):
            return ""

        meds = df_medicacion.copy()

        if "fecha_cambio" not in meds.columns:
            return ""

        meds["fecha_cambio"] = pd.to_datetime(meds["fecha_cambio"], errors="coerce")
        fecha_ref = pd.to_datetime(fecha_referencia, errors="coerce")

        meds = meds.dropna(subset=["fecha_cambio"])
        meds = meds[meds["fecha_cambio"] <= fecha_ref]

        if meds.empty:
            return ""

        meds = meds.sort_values("fecha_cambio", ascending=True)
        ultima = meds.iloc[-1]

        droga = str(ultima.get("droga", "")).strip()
        dosis = ultima.get("dosis")
        unidad = str(ultima.get("unidad", "")).strip()

        dosis_txt = ""
        if pd.notna(dosis):
            try:
                if float(dosis).is_integer():
                    dosis_txt = str(int(float(dosis)))
                else:
                    dosis_txt = str(round(float(dosis), 2))
            except Exception:
                dosis_txt = str(dosis).strip()

        if droga and dosis_txt and unidad:
            return f"{droga} {dosis_txt} {unidad}"
        elif droga and dosis_txt:
            return f"{droga} {dosis_txt}"
        elif droga:
            return droga
        else:
            return ""

    if df_eval is not None and not df_eval.empty:
        df_f = df_eval.copy()
        if "fecha" in df_f.columns:
            df_f["fecha"] = pd.to_datetime(df_f["fecha"], errors="coerce")

        for _, row in df_f.iterrows():
            fecha_row = row.get("fecha")
            prueba = str(row.get("prueba", "")).strip()
            valor = row.get("valor_medido")
            percentil = row.get("percentil")
            clasificacion = str(row.get("clasificacion", "")).strip()

            if prueba == "Prensión manual":
                resultado = f"{valor} kg" if pd.notna(valor) else ""
                evento = "Prensión manual"
            elif prueba == "Levantarse de la silla":
                resultado = f"{valor} rep" if pd.notna(valor) else ""
                evento = "Levantarse de la silla"
            elif prueba == "Caminata 6 minutos":
                resultado = f"{valor} m" if pd.notna(valor) else ""
                evento = "Caminata 6 minutos"
            else:
                resultado = str(valor) if pd.notna(valor) else ""
                evento = prueba

            filas.append({
                "Fecha": fecha_row,
                "Evento": evento,
                "Resultado": resultado,
                "Percentil": percentil,
                "Diagnostico": clasificacion,
                "Tratamiento": obtener_tratamiento_vigente(fecha_row)
            })

    if df_peso is not None and not df_peso.empty:
        df_p = df_peso.copy()
        if "fecha" in df_p.columns:
            df_p["fecha"] = pd.to_datetime(df_p["fecha"], errors="coerce")

        for _, row in df_p.iterrows():
            fecha_row = row.get("fecha")
            peso = row.get("peso_kg")
            imc = row.get("imc")

            peso_txt = f"{round(float(peso), 1)} kg" if pd.notna(peso) else "-"
            imc_txt = f"IMC {round(float(imc), 2)}" if pd.notna(imc) else "-"
            icc = row.get("icc")
            ica = row.get("ica")
            icc_txt = f"ICC {round(float(icc), 2)}" if pd.notna(icc) else ""
            ica_txt = f"ICA {round(float(ica), 2)}" if pd.notna(ica) else ""

            diagnostico = clasificar_imc(imc)[0] if pd.notna(imc) else ""

            partes_resultado = [peso_txt, imc_txt]
            if icc_txt:
                partes_resultado.append(icc_txt)
            if ica_txt:
                partes_resultado.append(ica_txt)

            filas.append({
                "Fecha": fecha_row,
                "Evento": "Peso / IMC",
                "Resultado": " / ".join(partes_resultado),
                "Percentil": None,
                "Diagnostico": diagnostico,
                "Tratamiento": obtener_tratamiento_vigente(fecha_row)
            })

    if df_inbody is not None and not df_inbody.empty:
        df_c = enriquecer_historial_corporal(df_inbody, str(sexo).strip().lower(), talla_m)
        if "fecha" in df_c.columns:
            df_c["fecha"] = pd.to_datetime(df_c["fecha"], errors="coerce")

        for _, row in df_c.iterrows():
            fecha_row = row.get("fecha")
            grasa = row.get("grasa_corporal_pct")
            musculo = row.get("masa_muscular_kg")
            diagnostico = str(row.get("diagnostico_corporal", "")).strip()

            partes = []
            if pd.notna(grasa):
                partes.append(f"{round(float(grasa), 1)}% grasa")
            if pd.notna(musculo):
                partes.append(f"{round(float(musculo), 1)} kg músculo")

            resultado = " / ".join(partes)

            filas.append({
                "Fecha": fecha_row,
                "Evento": "Composición corporal",
                "Resultado": resultado,
                "Percentil": None,
                "Diagnostico": diagnostico,
                "Tratamiento": obtener_tratamiento_vigente(fecha_row)
            })

    if df_medicacion is not None and not df_medicacion.empty:
        df_m = df_medicacion.copy()
        if "fecha_cambio" in df_m.columns:
            df_m["fecha_cambio"] = pd.to_datetime(df_m["fecha_cambio"], errors="coerce")

        for _, row in df_m.iterrows():
            fecha_row = row.get("fecha_cambio")
            droga = str(row.get("droga", "")).strip()
            dosis = row.get("dosis")
            unidad = str(row.get("unidad", "")).strip()
            frecuencia = str(row.get("frecuencia", "")).strip()
            via = str(row.get("via_administracion", "")).strip()
            estado = str(row.get("estado", "")).strip()

            dosis_txt = ""
            if pd.notna(dosis):
                try:
                    if float(dosis).is_integer():
                        dosis_txt = str(int(float(dosis)))
                    else:
                        dosis_txt = str(round(float(dosis), 2))
                except Exception:
                    dosis_txt = str(dosis).strip()

            resultado_partes = []
            if droga:
                resultado_partes.append(droga)
            if dosis_txt and unidad:
                resultado_partes.append(f"{dosis_txt} {unidad}")
            elif dosis_txt:
                resultado_partes.append(dosis_txt)
            if frecuencia:
                resultado_partes.append(f"{frecuencia}")
            if via:
                resultado_partes.append(via)

            filas.append({
                "Fecha": fecha_row,
                "Evento": "Tratamiento",
                "Resultado": " / ".join(resultado_partes),
                "Percentil": None,
                "Diagnostico": estado,
                "Tratamiento": obtener_tratamiento_vigente(fecha_row)
            })

    df_final = pd.DataFrame(filas)

    if not df_final.empty:
        df_final["Fecha"] = pd.to_datetime(df_final["Fecha"], errors="coerce")
        df_final = df_final.sort_values(["Fecha", "Evento"], ascending=[True, True]).reset_index(drop=True)

    return df_final


# =========================================================
# EXPORTACIÓN EXCEL
# =========================================================
def _formatear_hoja_excel(ws):
    if ws.max_row == 0 or ws.max_column == 0:
        return

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    fill_header = PatternFill(fill_type="solid", fgColor="1F4E78")
    font_header = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            try:
                cell_value = "" if cell.value is None else str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except Exception:
                pass

        ancho = min(max(max_length + 2, 12), 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def generar_tabla_estadistica(ficha, df_peso, df_inbody, df_eval, df_medicacion):
    from functools import reduce

    def normalizar_fecha_df(df, col_fecha):
        if df is None or df.empty:
            return df
        df = df.copy()
        df[col_fecha] = pd.to_datetime(df[col_fecha], errors="coerce").dt.normalize()
        df = df.dropna(subset=[col_fecha])
        return df

    if df_peso is not None and not df_peso.empty:
        peso = df_peso.copy()
        peso = normalizar_fecha_df(peso, "fecha")
        columnas_peso = ["fecha", "peso_kg", "imc", "cintura_cm", "cadera_cm", "icc", "ica"]
        columnas_peso = [c for c in columnas_peso if c in peso.columns]
        peso = peso[columnas_peso]
        peso["Riesgo_ICC"] = peso["icc"].apply(lambda x: clasificacion_icc(ficha.get("sexo"), x) if pd.notna(x) else "-") if "icc" in peso.columns else "-"
        peso["Riesgo_ICA"] = peso["ica"].apply(lambda x: clasificacion_ica(x) if pd.notna(x) else "-") if "ica" in peso.columns else "-"
        if "icc" in peso.columns and "ica" in peso.columns:
            peso["Clasificacion_Abdominal"] = peso.apply(
                lambda row: clasificacion_obesidad_abdominal(ficha.get("sexo"), row.get("icc"), row.get("ica")),
                axis=1
            )
        else:
            peso["Clasificacion_Abdominal"] = "-"
    else:
        peso = pd.DataFrame(columns=["fecha", "peso_kg", "imc", "cintura_cm", "cadera_cm", "icc", "ica", "Riesgo_ICC", "Riesgo_ICA", "Clasificacion_Abdominal"])

    if df_inbody is not None and not df_inbody.empty:
        inbody = df_inbody.copy()
        inbody = normalizar_fecha_df(inbody, "fecha")
        inbody = inbody[
            [
                "fecha",
                "grasa_corporal_pct",
                "masa_muscular_kg",
                "agua_corporal_pct",
                "grasa_visceral"
            ]
        ]
    else:
        inbody = pd.DataFrame(columns=[
            "fecha",
            "grasa_corporal_pct",
            "masa_muscular_kg",
            "agua_corporal_pct",
            "grasa_visceral"
        ])

    if df_eval is not None and not df_eval.empty:
        evals = df_eval.copy()
        evals = normalizar_fecha_df(evals, "fecha")

        prension = evals[evals["prueba"] == "Prensión manual"][["fecha", "valor_medido", "percentil"]].copy()
        prension.columns = ["fecha", "Prension_kg", "Prension_percentil"]

        silla = evals[evals["prueba"] == "Levantarse de la silla"][["fecha", "valor_medido", "percentil"]].copy()
        silla.columns = ["fecha", "SitToStand_rep", "SitToStand_percentil"]

        caminata = evals[evals["prueba"] == "Caminata 6 minutos"][["fecha", "valor_medido", "percentil"]].copy()
        caminata.columns = ["fecha", "Caminata6m_m", "Caminata6m_percentil"]
    else:
        prension = pd.DataFrame(columns=["fecha", "Prension_kg", "Prension_percentil"])
        silla = pd.DataFrame(columns=["fecha", "SitToStand_rep", "SitToStand_percentil"])
        caminata = pd.DataFrame(columns=["fecha", "Caminata6m_m", "Caminata6m_percentil"])

    if df_medicacion is not None and not df_medicacion.empty:
        med = df_medicacion.copy()
        med = normalizar_fecha_df(med, "fecha_cambio")
        med = med[
            [
                "fecha_cambio",
                "droga",
                "dosis",
                "unidad",
                "frecuencia",
                "via_administracion",
                "estado"
            ]
        ].copy()
        med.columns = [
            "fecha",
            "Droga",
            "Dosis",
            "Unidad",
            "Frecuencia",
            "Via",
            "Estado"
        ]
        med["fecha"] = pd.to_datetime(med["fecha"], errors="coerce").dt.normalize()
        med = med.dropna(subset=["fecha"])
    else:
        med = pd.DataFrame(columns=["fecha", "Droga", "Dosis", "Unidad", "Frecuencia", "Via", "Estado"])

    dfs = [peso, inbody, prension, silla, caminata, med]
    dfs_validos = [df for df in dfs if df is not None and not df.empty]

    if not dfs_validos:
        return pd.DataFrame()

    for i in range(len(dfs_validos)):
        dfs_validos[i] = dfs_validos[i].copy()
        dfs_validos[i]["fecha"] = pd.to_datetime(dfs_validos[i]["fecha"], errors="coerce").dt.normalize()

    df_final = reduce(lambda left, right: pd.merge(left, right, on="fecha", how="outer"), dfs_validos)

    df_final = df_final.sort_values("fecha").reset_index(drop=True)

    df_final.rename(
        columns={
            "fecha": "Fecha",
            "peso_kg": "Peso_kg",
            "imc": "IMC",
            "cintura_cm": "Cintura_cm",
            "cadera_cm": "Cadera_cm",
            "icc": "ICC",
            "ica": "ICA",
            "Riesgo_ICC": "Riesgo_ICC",
            "Riesgo_ICA": "Riesgo_ICA",
            "Clasificacion_Abdominal": "Clasificacion_Abdominal",
            "grasa_corporal_pct": "Grasa_pct",
            "masa_muscular_kg": "Musculo_kg",
            "agua_corporal_pct": "Agua_pct",
            "grasa_visceral": "Grasa_Visceral"
        },
        inplace=True
    )

    df_final.insert(0, "PacienteID_Ficha", ficha.get("id"))
    df_final.insert(0, "Paciente", ficha.get("nombre"))
    df_final.insert(2, "Sexo", ficha.get("sexo"))
    df_final.insert(3, "FechaNacimiento", ficha.get("fecha_nacimiento"))

    df_final["Fecha"] = pd.to_datetime(df_final["Fecha"], errors="coerce")
    df_final["FechaNacimiento"] = pd.to_datetime(df_final["FechaNacimiento"], errors="coerce")
    df_final["Edad"] = ((df_final["Fecha"] - df_final["FechaNacimiento"]).dt.days / 365.25).round(1)

    columnas_orden = [
        "Paciente",
        "PacienteID_Ficha",
        "Sexo",
        "FechaNacimiento",
        "Edad",
        "Fecha",
        "Peso_kg",
        "IMC",
        "Cintura_cm",
        "Cadera_cm",
        "ICC",
        "ICA",
        "Grasa_pct",
        "Musculo_kg",
        "Agua_pct",
        "Grasa_Visceral",
        "Prension_kg",
        "Prension_percentil",
        "SitToStand_rep",
        "SitToStand_percentil",
        "Caminata6m_m",
        "Caminata6m_percentil",
        "Droga",
        "Dosis",
        "Unidad",
        "Frecuencia",
        "Via",
        "Estado"
    ]

    columnas_presentes = [c for c in columnas_orden if c in df_final.columns]
    df_final = df_final[columnas_presentes]

    return df_final


def generar_excel_general(pacientes):
    output = BytesIO()
    tablas = []

    for p in pacientes:
        ficha_local = {
            "id": p.get("id"),
            "nombre": p.get("nombre"),
            "sexo": p.get("sexo"),
            "fecha_nacimiento": p.get("fecha_nacimiento"),
            "talla_m": p.get("talla_m")
        }

        paciente_id = p.get("id")

        df_peso_p = obtener_historial_peso(paciente_id)
        df_inbody_p = obtener_historial_inbody(paciente_id)
        df_eval_p = obtener_historial_paciente(paciente_id)
        df_medicacion_p = obtener_historial_medicacion(paciente_id)

        tabla_p = generar_tabla_estadistica(
            ficha=ficha_local,
            df_peso=df_peso_p,
            df_inbody=df_inbody_p,
            df_eval=df_eval_p,
            df_medicacion=df_medicacion_p
        )

        if tabla_p is not None and not tabla_p.empty:
            tablas.append(tabla_p)

    if tablas:
        df_estadistico = pd.concat(tablas, ignore_index=True)
        df_estadistico["Fecha"] = pd.to_datetime(df_estadistico["Fecha"], errors="coerce")
        df_estadistico = df_estadistico.sort_values(["Paciente", "Fecha"]).reset_index(drop=True)
    else:
        df_estadistico = pd.DataFrame(columns=[
            "Paciente",
            "PacienteID_Ficha",
            "Sexo",
            "FechaNacimiento",
            "Edad",
            "Fecha",
            "Peso_kg",
            "IMC",
            "Cintura_cm",
            "Cadera_cm",
            "ICC",
            "ICA",
            "Riesgo_ICC",
            "Riesgo_ICA",
            "Clasificacion_Abdominal",
            "Grasa_pct",
            "Musculo_kg",
            "Agua_pct",
            "Grasa_Visceral",
            "Prension_kg",
            "Prension_percentil",
            "SitToStand_rep",
            "SitToStand_percentil",
            "Caminata6m_m",
            "Caminata6m_percentil",
            "Droga",
            "Dosis",
            "Unidad",
            "Frecuencia",
            "Via",
            "Estado"
        ])

    df_estadistico_limpio = preparar_df_estadistico(preparar_df_exportacion(df_estadistico))
    df_longitudinal = preparar_dataset_longitudinal(df_estadistico_limpio)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_estadistico_limpio.to_excel(
            writer,
            sheet_name="Datos_Estadisticos",
            index=False
        )

        df_longitudinal.to_excel(
            writer,
            sheet_name="Dataset_Longitudinal",
            index=False
        )

        workbook = writer.book
        _formatear_hoja_excel(workbook["Datos_Estadisticos"])
        _formatear_hoja_excel(workbook["Dataset_Longitudinal"])

    output.seek(0)
    return output

# =========================================================
# EXPORTACIÓN PDF
# =========================================================
def _texto_seguro(valor):
    if pd.isna(valor):
        return "-"
    valor = str(valor).strip()
    return valor if valor else "-"


def _df_para_pdf(df):
    if df is None or df.empty:
        return pd.DataFrame()

    out = df.copy()

    for col in out.columns:
        if "fecha" in col.lower() or "created" in col.lower():
            out[col] = pd.to_datetime(out[col], errors="coerce").dt.strftime("%Y-%m-%d")

    return out.fillna("-")


def _valor_numerico_o_none(valor):
    if valor in [None, "", "-"]:
        return None
    try:
        if pd.isna(valor):
            return None
    except Exception:
        pass
    try:
        return float(valor)
    except (TypeError, ValueError):
        return None


def _formatear_numero_pdf(valor, decimales=2):
    numero = _valor_numerico_o_none(valor)
    if numero is None:
        return "-"
    return f"{numero:.{decimales}f}"


def _tabla_pdf_desde_df(df, columnas, titulos, anchos_cm, styles):
    if df is None or df.empty:
        return Paragraph("Sin datos.", styles["Normal"])

    df = df.copy()
    columnas_validas = [c for c in columnas if c in df.columns]
    titulos_validos = [titulos[i] for i, c in enumerate(columnas) if c in df.columns]
    anchos_validos = [anchos_cm[i] * cm for i, c in enumerate(columnas) if c in df.columns]

    data = [[Paragraph(f"<b>{t}</b>", styles["TablaHeader"]) for t in titulos_validos]]

    for _, row in df.iterrows():
        fila = []
        for c in columnas_validas:
            txt = _texto_seguro(row.get(c))
            fila.append(Paragraph(txt, styles["TablaBody"]))
        data.append(fila)

    tabla = Table(data, colWidths=anchos_validos, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7.5),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C9D2DF")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#F7FAFC")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return tabla
    
def header_pdf(canvas, doc):
    logo_path = "LogoPetratti.jpeg"

    try:
        canvas.drawImage(
            logo_path,
            doc.pagesize[0] - 7 * cm,
            doc.pagesize[1] - 3.2 * cm,
            width=6 * cm,
            height=2.4 * cm,
            preserveAspectRatio=True,
            mask="auto"
        )
    except Exception:
        pass

def generar_pdf_paciente(ficha, df_peso, df_inbody, df_eval, df_medicacion):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TituloMain", parent=styles["Heading1"], fontSize=18, leading=22, spaceAfter=10))
    styles.add(ParagraphStyle(name="SubTitulo", parent=styles["Heading2"], fontSize=12, leading=15, spaceBefore=8, spaceAfter=6))
    styles.add(ParagraphStyle(name="TablaHeader", parent=styles["Normal"], fontSize=8, textColor=colors.white))
    styles.add(ParagraphStyle(name="TablaBody", parent=styles["Normal"], fontSize=7.5, leading=9))
    styles.add(ParagraphStyle(name="Caja", parent=styles["Normal"], fontSize=9, leading=12))
    styles.add(ParagraphStyle(name="Firma", parent=styles["Normal"], fontSize=9.5, leading=12, spaceAfter=2))

    story = []

    story.append(Paragraph("Método Dra. Petratti", styles["TituloMain"]))
    story.append(Paragraph(f"<b>Paciente:</b> {_texto_seguro(ficha.get('nombre'))}", styles["Caja"]))
    story.append(Paragraph(f"<b>Sexo:</b> {_texto_seguro(ficha.get('sexo'))}", styles["Caja"]))
    story.append(Paragraph(f"<b>Talla:</b> {_texto_seguro(ficha.get('talla_m'))} m", styles["Caja"]))
    story.append(Paragraph(f"<b>Fecha del reporte:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Caja"]))
    story.append(Spacer(1, 0.25 * cm))

    df_peso_pdf = _df_para_pdf(df_peso)
    df_inbody_pdf = _df_para_pdf(df_inbody)
    df_eval_pdf = _df_para_pdf(df_eval)
    df_medicacion_pdf = _df_para_pdf(df_medicacion)

    df_analisis_pdf = generar_df_analisis_cientifico(
        ficha=ficha,
        df_peso=df_peso,
        df_inbody=df_inbody,
        df_eval=df_eval,
        df_medicacion=df_medicacion
    )
    df_analisis_pdf = _df_para_pdf(df_analisis_pdf)

    if df_inbody_pdf is not None and not df_inbody_pdf.empty:
        df_inbody_pdf = enriquecer_historial_corporal(
            df_inbody_pdf,
            str(ficha.get("sexo", "")).strip().lower(),
            ficha.get("talla_m")
        )
        df_inbody_pdf = _df_para_pdf(df_inbody_pdf)

    if df_inbody_pdf is not None and not df_inbody_pdf.empty:
        ultimo_corporal = df_inbody_pdf.copy()
        if "fecha" in ultimo_corporal.columns:
            ultimo_corporal["fecha_orden"] = pd.to_datetime(ultimo_corporal["fecha"], errors="coerce")
            ultimo_corporal = ultimo_corporal.sort_values("fecha_orden", ascending=False)
        ult = ultimo_corporal.iloc[0]

        story.append(Paragraph("Diagnóstico corporal actual", styles["SubTitulo"]))
        story.append(Paragraph(f"<b>Diagnóstico:</b> {_texto_seguro(ult.get('diagnostico_corporal'))}", styles["Caja"]))
        story.append(Paragraph(f"<b>Sugerencia:</b> {_texto_seguro(ult.get('sugerencia_corporal'))}", styles["Caja"]))
        story.append(Paragraph(f"<b>Motivos:</b> {_texto_seguro(ult.get('motivos_corporal'))}", styles["Caja"]))
        story.append(Spacer(1, 0.25 * cm))

    story.append(Paragraph("Historial de peso e IMC", styles["SubTitulo"]))
    story.append(_tabla_pdf_desde_df(
        df_peso_pdf,
        columnas=["fecha", "peso_kg", "imc", "cintura_cm", "cadera_cm", "icc", "ica"],
        titulos=["Fecha", "Peso (kg)", "IMC", "Cintura", "Cadera", "ICC", "ICA"],
        anchos_cm=[2.4, 2.1, 1.6, 2.0, 2.0, 1.5, 1.5],
        styles=styles
    ))
    story.append(Spacer(1, 0.2 * cm))

    ultimo_peso_pdf = obtener_ultimo_registro(df_peso_pdf, "fecha")
    if ultimo_peso_pdf is not None:
        icc_pdf = _valor_numerico_o_none(ultimo_peso_pdf.get("icc"))
        ica_pdf = _valor_numerico_o_none(ultimo_peso_pdf.get("ica"))

        clasif_icc_pdf = clasificacion_icc(ficha.get("sexo"), icc_pdf)
        clasif_ica_pdf = clasificacion_ica(ica_pdf)
        clasif_abdominal_pdf = clasificacion_obesidad_abdominal(ficha.get("sexo"), icc_pdf, ica_pdf)

        story.append(Paragraph("Clasificación abdominal automática", styles["SubTitulo"]))
        story.append(Paragraph(
            f"<b>ICC:</b> {_formatear_numero_pdf(icc_pdf, 2)} — {_texto_seguro(clasif_icc_pdf)}",
            styles["Caja"]
        ))
        story.append(Paragraph(
            f"<b>ICA:</b> {_formatear_numero_pdf(ica_pdf, 2)} — {_texto_seguro(clasif_ica_pdf)}",
            styles["Caja"]
        ))
        story.append(Paragraph(
            f"<b>Conclusión abdominal:</b> {_texto_seguro(clasif_abdominal_pdf)}",
            styles["Caja"]
        ))
        story.append(Spacer(1, 0.25 * cm))
    else:
        story.append(Spacer(1, 0.35 * cm))

    story.append(Paragraph("Historial de composición corporal", styles["SubTitulo"]))
    story.append(_tabla_pdf_desde_df(
        df_inbody_pdf,
        columnas=[
            "fecha",
            "peso_kg",
            "imc",
            "grasa_corporal_pct",
            "masa_muscular_kg",
            "agua_corporal_pct",
            "grasa_visceral",
            "diagnostico_corporal"
        ],
        titulos=[
            "Fecha",
            "Peso",
            "IMC",
            "% Grasa",
            "Músculo kg",
            "% Agua",
            "Visceral",
            "Diagnóstico"
        ],
        anchos_cm=[2.8, 2.1, 1.7, 2.2, 2.4, 2.1, 1.8, 4.1],
        styles=styles
    ))
    story.append(Spacer(1, 0.35 * cm))

    story.append(Paragraph("Evaluaciones funcionales", styles["SubTitulo"]))
    story.append(_tabla_pdf_desde_df(
        df_eval_pdf,
        columnas=["fecha", "prueba", "valor_medido", "percentil", "clasificacion"],
        titulos=["Fecha", "Prueba", "Valor", "Percentil", "Clasificación"],
        anchos_cm=[2.8, 5.7, 2.2, 2.0, 3.5],
        styles=styles
    ))
    story.append(Spacer(1, 0.35 * cm))

    story.append(Paragraph("Historial de medicación", styles["SubTitulo"]))
    story.append(_tabla_pdf_desde_df(
        df_medicacion_pdf,
        columnas=["fecha_cambio", "droga", "dosis", "unidad", "frecuencia", "via_administracion", "estado", "observaciones"],
        titulos=["Fecha", "Droga", "Dosis", "Unidad", "Frecuencia", "Vía", "Estado", "Observaciones"],
        anchos_cm=[2.2, 3.3, 1.6, 1.5, 2.3, 2.1, 1.8, 3.2],
        styles=styles
    ))
    story.append(Spacer(1, 0.35 * cm))

    story.append(Paragraph("Análisis integrado", styles["SubTitulo"]))
    story.append(_tabla_pdf_desde_df(
        df_analisis_pdf,
        columnas=["Fecha", "Evento", "Resultado", "Percentil", "Diagnostico", "Tratamiento"],
        titulos=["Fecha", "Evento", "Resultado", "Percentil", "Diagnóstico", "Tratamiento"],
        anchos_cm=[2.0, 3.4, 4.2, 1.5, 2.8, 3.6],
        styles=styles
    ))

    informe_pdf = generar_informe_integrado_paciente(
        ficha=ficha,
        df_peso=df_peso,
        df_inbody=df_inbody,
        df_eval=df_eval,
        df_medicacion=df_medicacion
    )
    percentiles_pdf = informe_pdf.get("percentiles_funcionales", {})
    p_caminata_pdf = percentiles_pdf.get("Caminata 6 minutos")
    p_prension_pdf = percentiles_pdf.get("Prensión manual")
    p_silla_pdf = percentiles_pdf.get("Levantarse de la silla")

    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph("Resumen funcional integrado", styles["SubTitulo"]))
    story.append(Paragraph(
        f"<b>Caminata 6 minutos:</b> {('P' + str(p_caminata_pdf)) if p_caminata_pdf is not None else '-'}",
        styles["Caja"]
    ))
    story.append(Paragraph(
        f"<b>Prensión manual:</b> {('P' + str(p_prension_pdf)) if p_prension_pdf is not None else '-'}",
        styles["Caja"]
    ))
    story.append(Paragraph(
        f"<b>Levantarse de la silla:</b> {('P' + str(p_silla_pdf)) if p_silla_pdf is not None else '-'}",
        styles["Caja"]
    ))
    story.append(Paragraph(
        f"<b>Comentario clínico unificado:</b> {informe_pdf.get('comentario_unificado', '-')}",
        styles["Caja"]
    ))
    story.append(Paragraph(
        f"<b>Recomendación unificada:</b> {informe_pdf.get('recomendacion_final', '-')}",
        styles["Caja"]
    ))

    story.append(Spacer(1, 0.55 * cm))
    story.append(Paragraph("<font color='#B0B7C3'>______________________________________________</font>", styles["Caja"]))
    story.append(Spacer(1, 0.12 * cm))
    story.append(Paragraph("<b>Dra. Cristina B. Petratti</b>", styles["Firma"]))
    story.append(Paragraph("Médico de Familia – Especialista en Obesidad y Nutrición", styles["Firma"]))
    story.append(Paragraph("Nº Colegiada: 032880978", styles["Firma"]))
    story.append(Paragraph("Método Dra. Petratti", styles["Firma"]))
    story.append(Paragraph("www.cristinapetratti.com", styles["Firma"]))
    story.append(Paragraph("Instagram: @crispetratti", styles["Firma"]))

    doc.build(
    story,
    onFirstPage=header_pdf,
    onLaterPages=header_pdf
    )
    buffer.seek(0)
    return buffer


# =========================================================
# UTILIDADES CLÍNICAS - FUNCIONALES
# =========================================================
def clasificar_percentil(percentil):
    if percentil is None:
        return "Sin clasificar"
    if percentil < 10:
        return "Muy bajo"
    if percentil < 25:
        return "Bajo"
    if percentil < 50:
        return "Ligeramente bajo"
    if percentil < 75:
        return "Normal"
    if percentil < 90:
        return "Bueno"
    return "Muy bueno"


def color_clasificacion(clasificacion):
    colores = {
        "Muy bajo": "#d32f2f",
        "Bajo": "#f57c00",
        "Ligeramente bajo": "#fbc02d",
        "Normal": "#388e3c",
        "Bueno": "#1976d2",
        "Muy bueno": "#00796b",
        "Sin clasificar": "#757575"
    }
    return colores.get(clasificacion, "#757575")


def rango_percentilar(percentil):
    if percentil is None:
        return "Sin rango"
    if percentil < 3:
        return "Menor a P3"
    if percentil < 10:
        return "Entre P3 y P10"
    if percentil < 25:
        return "Entre P10 y P25"
    if percentil < 50:
        return "Entre P25 y P50"
    if percentil < 75:
        return "Entre P50 y P75"
    if percentil < 90:
        return "Entre P75 y P90"
    if percentil < 97:
        return "Entre P90 y P97"
    return "Mayor a P97"


def interpretacion_clinica(clasificacion):
    textos = {
        "Muy bajo": "Resultado marcadamente por debajo del rango funcional esperado.",
        "Bajo": "Resultado por debajo del rango funcional esperado.",
        "Ligeramente bajo": "Resultado levemente inferior al rango esperado.",
        "Normal": "Resultado dentro del rango funcional esperado.",
        "Bueno": "Resultado superior al promedio esperado.",
        "Muy bueno": "Resultado claramente superior al rango esperado."
    }
    return textos.get(clasificacion, "Sin interpretación disponible.")


def interpolar_percentil(valor_medido, tabla_percentiles):
    puntos = sorted(tabla_percentiles.items(), key=lambda x: x[1])

    if valor_medido < puntos[0][1]:
        return float(puntos[0][0])

    if valor_medido > puntos[-1][1]:
        return float(puntos[-1][0])

    for i in range(len(puntos) - 1):
        p1, v1 = puntos[i]
        p2, v2 = puntos[i + 1]

        if v1 <= valor_medido <= v2:
            if v2 == v1:
                return float(p1)
            return float(p1 + (valor_medido - v1) * (p2 - p1) / (v2 - v1))

    return None


def grupo_edad_prension(edad):
    edad = int(edad)
    if edad >= 100:
        return "+100"
    inicio = max(20, min(95, 20 + ((edad - 20) // 5) * 5))
    fin = inicio + 4
    return f"{inicio}-{fin}"


def grupo_edad_silla(edad):
    edad = int(edad)
    if edad >= 85:
        return "+84"
    if 65 <= edad <= 69:
        return "65-69"
    if 70 <= edad <= 74:
        return "70-74"
    if 75 <= edad <= 79:
        return "75-79"
    if 80 <= edad <= 84:
        return "80-84"
    return None


def obtener_altura_referencia_caminata(talla_m):
    if talla_m is None:
        return None
    try:
        altura_cm = int(round(float(talla_m) * 100))
    except Exception:
        return None

    alturas_validas = sorted(TABLA_CAMINATA_6M.keys())
    if altura_cm < min(alturas_validas) or altura_cm > max(alturas_validas):
        return None

    return min(alturas_validas, key=lambda x: abs(x - altura_cm))


def calcular_resultado(prueba, sexo, edad, altura, valor_medido):
    sexo = str(sexo).strip()
    edad = int(edad)
    valor_medido = float(valor_medido)

    if valor_medido <= 0:
        return None, "Sin clasificar", "-", "-", "-"

    if prueba == "Caminata 6 minutos":
        alturas_validas = sorted(TABLA_CAMINATA_6M.keys())
        edades_validas = sorted(next(iter(TABLA_CAMINATA_6M.values())).keys())

        if altura is None:
            return None, "Sin clasificar", "-", "-", "Altura fuera del rango validado por la tabla"

        altura = int(altura)
        if altura < min(alturas_validas) or altura > max(alturas_validas):
            return None, "Sin clasificar", "-", "-", "Altura fuera del rango validado por la tabla"

        if edad < min(edades_validas) or edad > max(edades_validas):
            return None, "Sin clasificar", "-", "-", "Edad fuera del rango validado por la tabla"

        altura_ref = min(alturas_validas, key=lambda x: abs(x - altura))
        edad_ref = min(edades_validas, key=lambda x: abs(x - edad))

        percentiles = TABLA_CAMINATA_6M[altura_ref][edad_ref]
        percentil_estimado = interpolar_percentil(valor_medido, percentiles)

        if percentil_estimado is None:
            return None, "Sin clasificar", "-", "-", "-"

        percentil_estimado = round(percentil_estimado, 1)
        clasificacion = clasificar_percentil(percentil_estimado)
        referencia_p50 = percentiles[50]

        return (
            percentil_estimado,
            clasificacion,
            f"{referencia_p50} m",
            f"Altura ref.: {altura_ref} cm",
            f"Edad ref.: {edad_ref} años"
        )

    if prueba == "Prensión manual":
        if edad < 20 or edad > 100:
            return None, "Sin clasificar", "-", "-", "Edad fuera del rango validado por la tabla"

        grupo = grupo_edad_prension(edad)
        percentiles = TABLA_PRENSION[sexo][grupo]
        percentil_estimado = interpolar_percentil(valor_medido, percentiles)

        if percentil_estimado is None:
            return None, "Sin clasificar", "-", "-", "-"

        percentil_estimado = round(percentil_estimado, 1)
        clasificacion = clasificar_percentil(percentil_estimado)
        referencia_p50 = percentiles[50]

        return (
            percentil_estimado,
            clasificacion,
            f"{referencia_p50} kg",
            "-",
            f"Grupo etario: {grupo}"
        )

    if prueba == "Levantarse de la silla":
        grupo = grupo_edad_silla(edad)

        if grupo is None:
            return None, "Sin clasificar", "-", "-", "Esta prueba percentilar está validada solo para pacientes de 65 años o más"

        percentiles = TABLA_SILLA[sexo][grupo]
        percentil_estimado = interpolar_percentil(valor_medido, percentiles)

        if percentil_estimado is None:
            return None, "Sin clasificar", "-", "-", "-"

        percentil_estimado = round(percentil_estimado, 1)
        clasificacion = clasificar_percentil(percentil_estimado)
        referencia_p50 = percentiles[50]

        return (
            percentil_estimado,
            clasificacion,
            f"{referencia_p50} rep",
            "-",
            f"Grupo etario: {grupo}"
        )

    return None, "Sin clasificar", "-", "-", "-"



def obtener_ultimo_registro(df, columna_fecha="fecha"):
    if df is None or df.empty:
        return None

    df_tmp = df.copy()

    if columna_fecha in df_tmp.columns:
        df_tmp[columna_fecha] = pd.to_datetime(df_tmp[columna_fecha], errors="coerce")

    columnas_sort = []
    ascending_sort = []

    if columna_fecha in df_tmp.columns:
        columnas_sort.append(columna_fecha)
        ascending_sort.append(False)

    if "created_at" in df_tmp.columns:
        df_tmp["created_at"] = pd.to_datetime(df_tmp["created_at"], errors="coerce")
        columnas_sort.append("created_at")
        ascending_sort.append(False)

    if "id" in df_tmp.columns:
        columnas_sort.append("id")
        ascending_sort.append(False)

    if columnas_sort:
        df_tmp = df_tmp.sort_values(columnas_sort, ascending=ascending_sort)

    if df_tmp.empty:
        return None

    return df_tmp.iloc[0]


def color_estado_global_informe(estado):
    mapa = {
        "Perfil conservado": ("#2e7d32", "#ffffff"),
        "Riesgo funcional leve": ("#f9a825", "#1f1f1f"),
        "Riesgo funcional moderado": ("#ef6c00", "#ffffff"),
        "Riesgo funcional alto": ("#c62828", "#ffffff"),
        "Riesgo cardiometabólico": ("#ad1457", "#ffffff"),
        "Riesgo combinado": ("#6a1b9a", "#ffffff"),
        "Sin datos suficientes": ("#757575", "#ffffff")
    }
    return mapa.get(estado, ("#757575", "#ffffff"))


def generar_informe_integrado_paciente(ficha, df_peso, df_inbody, df_eval, df_medicacion):
    filas_resumen = []

    nombre = ficha.get("nombre", "-")
    sexo = str(ficha.get("sexo", "")).strip().lower()
    talla_m = ficha.get("talla_m")

    ultimo_peso = obtener_ultimo_registro(df_peso, "fecha")
    ultimo_inbody = None
    if df_inbody is not None and not df_inbody.empty and talla_m is not None:
        df_inbody_enriquecido = enriquecer_historial_corporal(df_inbody, sexo, talla_m)
        ultimo_inbody = obtener_ultimo_registro(df_inbody_enriquecido, "fecha")
    else:
        df_inbody_enriquecido = pd.DataFrame()

    df_eval_tmp = df_eval.copy() if df_eval is not None else pd.DataFrame()
    if not df_eval_tmp.empty and "fecha" in df_eval_tmp.columns:
        df_eval_tmp["fecha"] = pd.to_datetime(df_eval_tmp["fecha"], errors="coerce")
        df_eval_tmp = df_eval_tmp.sort_values(["prueba", "fecha"], ascending=[True, False])

    ultimos_funcionales = {}
    if not df_eval_tmp.empty:
        for prueba in ["Caminata 6 minutos", "Prensión manual", "Levantarse de la silla"]:
            df_prueba = df_eval_tmp[df_eval_tmp["prueba"].astype(str).str.strip() == prueba].copy()
            if not df_prueba.empty:
                ultimos_funcionales[prueba] = df_prueba.iloc[0]

    clasif_abdominal = "Sin clasificar"

    if ultimo_peso is not None:
        filas_resumen.append({
            "Área": "Corporal",
            "Variable": "IMC",
            "Valor": round(float(ultimo_peso.get("imc")), 2) if pd.notna(ultimo_peso.get("imc")) else "-",
            "Percentil": "-",
            "Clasificación": clasificar_imc(float(ultimo_peso.get("imc")))[0] if pd.notna(ultimo_peso.get("imc")) else "-"
        })
        filas_resumen.append({
            "Área": "Corporal",
            "Variable": "Índice cintura-cadera",
            "Valor": round(float(ultimo_peso.get("icc")), 2) if pd.notna(ultimo_peso.get("icc")) else "-",
            "Percentil": "-",
            "Clasificación": clasificacion_icc(sexo, ultimo_peso.get("icc")) if pd.notna(ultimo_peso.get("icc")) else "-"
        })
        filas_resumen.append({
            "Área": "Corporal",
            "Variable": "Índice cintura-altura",
            "Valor": round(float(ultimo_peso.get("ica")), 2) if pd.notna(ultimo_peso.get("ica")) else "-",
            "Percentil": "-",
            "Clasificación": clasificacion_ica(ultimo_peso.get("ica")) if pd.notna(ultimo_peso.get("ica")) else "-"
        })
        clasif_abdominal = clasificacion_obesidad_abdominal(sexo, ultimo_peso.get("icc"), ultimo_peso.get("ica"))
        filas_resumen.append({
            "Área": "Corporal",
            "Variable": "Obesidad abdominal",
            "Valor": "ICC + ICA",
            "Percentil": "-",
            "Clasificación": clasif_abdominal
        })

    if ultimo_inbody is not None:
        filas_resumen.append({
            "Área": "Corporal",
            "Variable": "Grasa corporal",
            "Valor": f"{round(float(ultimo_inbody.get('grasa_corporal_pct')), 1)} %" if pd.notna(ultimo_inbody.get("grasa_corporal_pct")) else "-",
            "Percentil": "-",
            "Clasificación": str(ultimo_inbody.get("clasif_grasa", "-"))
        })
        filas_resumen.append({
            "Área": "Corporal",
            "Variable": "Músculo relativo",
            "Valor": f"{round(float(ultimo_inbody.get('musculo_rel_pct')), 1)} %" if pd.notna(ultimo_inbody.get("musculo_rel_pct")) else "-",
            "Percentil": "-",
            "Clasificación": str(ultimo_inbody.get("clasif_musculo", "-"))
        })
        filas_resumen.append({
            "Área": "Corporal",
            "Variable": "Grasa visceral",
            "Valor": round(float(ultimo_inbody.get("grasa_visceral")), 1) if pd.notna(ultimo_inbody.get("grasa_visceral")) else "-",
            "Percentil": "-",
            "Clasificación": str(ultimo_inbody.get("clasif_visceral", "-"))
        })

    for prueba, etiqueta_valor in {
        "Caminata 6 minutos": "m",
        "Prensión manual": "kg",
        "Levantarse de la silla": "rep"
    }.items():
        row = ultimos_funcionales.get(prueba)
        if row is not None:
            valor = row.get("valor_medido")
            percentil = row.get("percentil")
            clasificacion = row.get("clasificacion", "-")

            valor_fmt = "-"
            if pd.notna(valor):
                if etiqueta_valor == "rep":
                    valor_fmt = f"{int(round(float(valor)))} {etiqueta_valor}"
                else:
                    valor_fmt = f"{round(float(valor), 1)} {etiqueta_valor}"

            filas_resumen.append({
                "Área": "Funcional",
                "Variable": prueba,
                "Valor": valor_fmt,
                "Percentil": f"P{round(float(percentil), 1)}" if pd.notna(percentil) else "-",
                "Clasificación": str(clasificacion)
            })

    df_resumen = pd.DataFrame(filas_resumen)

    # Normalizar columnas mixtas para evitar errores de PyArrow/Streamlit
    # cuando una misma columna contiene números y textos (por ejemplo "ICC + ICA").
    if not df_resumen.empty:
        for col in ["Área", "Variable", "Valor", "Percentil", "Clasificación"]:
            if col in df_resumen.columns:
                df_resumen[col] = df_resumen[col].astype(str)

    percentiles_validos = []
    clasificaciones_funcionales = []
    percentiles_funcionales = {
        "Caminata 6 minutos": None,
        "Prensión manual": None,
        "Levantarse de la silla": None
    }

    for prueba, row in ultimos_funcionales.items():
        if pd.notna(row.get("percentil")):
            percentil_valor = float(row.get("percentil"))
            percentiles_validos.append(percentil_valor)
            percentiles_funcionales[prueba] = round(percentil_valor, 1)
        if pd.notna(row.get("clasificacion")):
            clasificaciones_funcionales.append(str(row.get("clasificacion")))

    peor_percentil = min(percentiles_validos) if percentiles_validos else None
    promedio_percentil = round(sum(percentiles_validos) / len(percentiles_validos), 1) if percentiles_validos else None

    cant_muy_bajo = sum(1 for x in percentiles_validos if x < 10)
    cant_bajo = sum(1 for x in percentiles_validos if 10 <= x < 25)

    estado_corporal = str(ultimo_inbody.get("diagnostico_corporal")) if ultimo_inbody is not None and pd.notna(ultimo_inbody.get("diagnostico_corporal")) else "Sin datos"
    recomendacion_corporal = str(ultimo_inbody.get("sugerencia_corporal")) if ultimo_inbody is not None and pd.notna(ultimo_inbody.get("sugerencia_corporal")) else ""

    if not filas_resumen:
        estado_global = "Sin datos suficientes"
    elif clasif_abdominal in ["Obesidad abdominal / riesgo alto", "Obesidad abdominal / riesgo aumentado"] and (cant_muy_bajo >= 1 or cant_bajo >= 2):
        estado_global = "Riesgo combinado"
    elif estado_corporal in ["Riesgo cardiometabólico", "Riesgo cardiometabólico moderado"] and (cant_muy_bajo >= 1 or cant_bajo >= 2):
        estado_global = "Riesgo combinado"
    elif clasif_abdominal in ["Obesidad abdominal / riesgo alto", "Obesidad abdominal / riesgo aumentado", "Riesgo abdominal aumentado"]:
        estado_global = "Riesgo cardiometabólico"
    elif estado_corporal in ["Riesgo cardiometabólico", "Riesgo cardiometabólico moderado", "Obesidad"]:
        estado_global = "Riesgo cardiometabólico"
    elif cant_muy_bajo >= 2 or (cant_muy_bajo >= 1 and cant_bajo >= 1):
        estado_global = "Riesgo funcional alto"
    elif cant_muy_bajo >= 1 or cant_bajo >= 2:
        estado_global = "Riesgo funcional moderado"
    elif cant_bajo == 1 or estado_corporal in ["Sobrepeso", "Riesgo sarcopénico", "Bajo peso"]:
        estado_global = "Riesgo funcional leve"
    else:
        estado_global = "Perfil conservado"

    partes_comentario = [f"Paciente {nombre}."]

    resumen_funcional = []
    if percentiles_funcionales.get("Caminata 6 minutos") is not None:
        resumen_funcional.append(f"Caminata 6 minutos: P{percentiles_funcionales['Caminata 6 minutos']}")
    if percentiles_funcionales.get("Prensión manual") is not None:
        resumen_funcional.append(f"Prensión manual: P{percentiles_funcionales['Prensión manual']}")
    if percentiles_funcionales.get("Levantarse de la silla") is not None:
        resumen_funcional.append(f"Levantarse de la silla: P{percentiles_funcionales['Levantarse de la silla']}")

    if resumen_funcional:
        partes_comentario.append("Percentiles funcionales recientes: " + "; ".join(resumen_funcional) + ".")

    if estado_corporal != "Sin datos":
        partes_comentario.append(f"Estado corporal actual: {estado_corporal}.")
    if clasif_abdominal != "Sin clasificar":
        partes_comentario.append(f"Clasificación abdominal: {clasif_abdominal}.")
    if cant_muy_bajo >= 1:
        partes_comentario.append("Hay al menos una prueba funcional en rango muy bajo.")
    elif cant_bajo >= 1:
        partes_comentario.append("Hay pruebas funcionales por debajo del rango esperado.")
    else:
        if percentiles_validos:
            partes_comentario.append("Las pruebas funcionales recientes no muestran descensos relevantes.")
    comentario_unificado = " ".join(partes_comentario).strip()

    recomendaciones = []
    if estado_global in ["Riesgo funcional alto", "Riesgo funcional moderado", "Riesgo combinado"]:
        recomendaciones.append("Priorizar entrenamiento de fuerza y capacidad funcional.")
    if estado_global in ["Riesgo cardiometabólico", "Riesgo combinado"] or estado_corporal in ["Obesidad", "Sobrepeso", "Riesgo cardiometabólico", "Riesgo cardiometabólico moderado"] or clasif_abdominal in ["Obesidad abdominal / riesgo alto", "Obesidad abdominal / riesgo aumentado", "Riesgo abdominal aumentado"]:
        recomendaciones.append("Controlar adiposidad abdominal y riesgo cardiometabólico.")
    if peor_percentil is not None:
        if peor_percentil < 10:
            recomendaciones.append("Repetir evaluación funcional en 30 días.")
        elif peor_percentil < 25:
            recomendaciones.append("Repetir evaluación funcional en 45-60 días.")
    if recomendacion_corporal:
        recomendaciones.append(recomendacion_corporal)

    recomendaciones = list(dict.fromkeys([r for r in recomendaciones if r]))
    recomendacion_final = " ".join(recomendaciones) if recomendaciones else "Mantener seguimiento periódico y repetir controles según evolución."

    return {
        "estado_global": estado_global,
        "estado_corporal": estado_corporal,
        "peor_percentil": peor_percentil,
        "promedio_percentil": promedio_percentil,
        "percentiles_funcionales": percentiles_funcionales,
        "comentario_unificado": comentario_unificado,
        "recomendacion_final": recomendacion_final,
        "tabla_resumen": df_resumen
    }

# =========================================================
# UI
# =========================================================
pacientes = obtener_pacientes()

if not pacientes:
    st.warning("No hay pacientes cargados.")
    st.stop()

st.title("Método Dra. Petratti")
st.caption("Evaluación clínica de condición física y riesgo cardiometabólico")

col_alta_btn_1, col_alta_btn_2 = st.columns([1, 5])

with col_alta_btn_1:
    if st.button("➕ Nuevo paciente", key="btn_mostrar_form_nuevo_paciente"):
        st.session_state["mostrar_form_nuevo_paciente"] = True

with col_alta_btn_2:
    if st.session_state["mostrar_form_nuevo_paciente"]:
        st.info("Completá los datos y guardá. Después vamos a hacer que se cierre solo al guardar.")

if st.session_state["mostrar_form_nuevo_paciente"]:
    with st.expander("➕ Nuevo paciente", expanded=True):
        nuevo_nombre = st.text_input("Nombre del nuevo paciente", key="nuevo_nombre_alta")
        nuevo_sexo = st.selectbox("Sexo del nuevo paciente", ["hombre", "mujer"], key="nuevo_sexo_alta")
        nueva_fecha_nacimiento = st.date_input(
            "Fecha de nacimiento",
            value=st.session_state.get("nueva_fecha_nacimiento_alta", date(1970, 1, 1)),
            min_value=date(1920, 1, 1),
            max_value=date.today(),
            format="DD/MM/YYYY",
            key="nueva_fecha_nacimiento_alta"
        )
        nueva_talla = st.number_input(
            "Talla (m)",
            min_value=0.50,
            max_value=2.50,
            step=0.01,
            format="%.2f",
            key="nueva_talla_alta"
        )

        col_guardar_paciente, col_cancelar_paciente = st.columns(2)

        with col_guardar_paciente:
            if st.button("Guardar paciente", key="btn_guardar_paciente"):
                if not nuevo_nombre.strip():
                    st.warning("Ingresá el nombre del paciente.")
                else:
                    try:
                        resp_nuevo_paciente = guardar_paciente(
                            nombre=nuevo_nombre,
                            sexo=nuevo_sexo,
                            fecha_nacimiento=nueva_fecha_nacimiento,
                            talla_m=nueva_talla
                        )

                        nuevo_paciente_id = None
                        if getattr(resp_nuevo_paciente, "data", None):
                            try:
                                nuevo_paciente_id = int(resp_nuevo_paciente.data[0]["id"])
                            except Exception:
                                nuevo_paciente_id = None

                        if nuevo_paciente_id is None:
                            paciente_recien_creado = obtener_paciente_por_nombre(nuevo_nombre)
                            if paciente_recien_creado is not None:
                                nuevo_paciente_id = paciente_recien_creado.get("id")

                        st.session_state["mostrar_form_nuevo_paciente"] = False
                        st.session_state["mostrar_form_editar_paciente"] = False
                        st.session_state["busqueda_paciente"] = ""
                        st.session_state["paciente_id_seleccionado"] = nuevo_paciente_id
                        st.session_state["paciente_nombre_pendiente"] = str(nuevo_nombre).strip().lower()
                        st.session_state["paciente_cargado_id"] = None

                        resetear_pruebas_funcionales()

                        st.success("Paciente agregado correctamente.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error al guardar paciente: {e}")

        with col_cancelar_paciente:
            if st.button("Cancelar", key="btn_cancelar_nuevo_paciente"):
                st.session_state["mostrar_form_nuevo_paciente"] = False
                st.rerun()

# =========================================================
# ENCABEZADO
# =========================================================
top1, top2, top3, top4 = st.columns([2, 1, 1, 1])

with top1:

    if st.session_state.get("limpiar_busqueda_pendiente", False):
        st.session_state["busqueda_paciente"] = ""
        st.session_state["limpiar_busqueda_pendiente"] = False

    busqueda_paciente = st.text_input(
        "Buscar paciente",
        placeholder="Escribí nombre o parte del nombre...",
        key="busqueda_paciente"
    )

    busqueda_normalizada = busqueda_paciente.strip().lower()

    if busqueda_normalizada:
        pacientes_filtrados = [
            p for p in pacientes
            if busqueda_normalizada in str(p.get("nombre", "")).strip().lower()
        ]
    else:
        pacientes_filtrados = pacientes

    if not pacientes_filtrados:
        st.warning("No se encontraron pacientes con esa búsqueda.")
        st.stop()

    paciente_id_preseleccionado = st.session_state.get("paciente_id_seleccionado")
    paciente_nombre_pendiente = st.session_state.get("paciente_nombre_pendiente")

    etiquetas = []
    mapa_etiqueta_id = {}

    for p in pacientes_filtrados:
        etiqueta = f"{p['nombre']} | ID {p['id']}"

        if p.get("fecha_nacimiento"):
            etiqueta += f" | Nac: {p['fecha_nacimiento']}"

        etiquetas.append(etiqueta)
        mapa_etiqueta_id[etiqueta] = p["id"]

    indice_default = 0
    if paciente_id_preseleccionado is not None:
        for i, p in enumerate(pacientes_filtrados):
            if p.get("id") == paciente_id_preseleccionado:
                indice_default = i
                break
    elif paciente_nombre_pendiente:
        for i, p in enumerate(pacientes_filtrados):
            if str(p.get("nombre", "")).strip().lower() == paciente_nombre_pendiente:
                indice_default = i
                break

    etiqueta_preseleccionada = etiquetas[indice_default] if etiquetas else None

    if st.session_state.get("selector_paciente") not in etiquetas and etiqueta_preseleccionada is not None:
        st.session_state["selector_paciente"] = etiqueta_preseleccionada

    if paciente_nombre_pendiente:
        for etiqueta in etiquetas:
            if paciente_nombre_pendiente == str(etiqueta).split(" | ID ")[0].strip().lower():
                st.session_state["selector_paciente"] = etiqueta
                st.session_state["paciente_nombre_pendiente"] = None
                break

    seleccion = st.selectbox(
        "Seleccionar paciente",
        etiquetas,
        key="selector_paciente"
    )

    paciente_id = mapa_etiqueta_id[seleccion]
    paciente_actual = next(p for p in pacientes_filtrados if p["id"] == paciente_id)
    paciente_nombre = paciente_actual["nombre"]

    st.session_state["paciente_id_seleccionado"] = paciente_id

    if st.button("✏️ Editar ficha", key=f"btn_mostrar_editar_paciente_{paciente_id}"):
        st.session_state["mostrar_form_editar_paciente"] = True

with top4:
    st.markdown("###")

    confirmar_eliminar = st.checkbox(
        "Confirmar borrado",
        key=f"confirmar_borrado_{paciente_id}"
    )

    if st.button("Eliminar paciente", key=f"btn_eliminar_paciente_{paciente_id}"):

        if not confirmar_eliminar:
            st.warning("Marcá la confirmación antes de eliminar.")

        else:
            try:
                eliminar_paciente(paciente_id)

                st.session_state["paciente_id_seleccionado"] = None
                st.session_state["paciente_nombre_pendiente"] = None
                st.session_state["paciente_cargado_id"] = None
                st.session_state["mostrar_form_nuevo_paciente"] = False
                st.session_state["limpiar_busqueda_pendiente"] = True

                st.success("Paciente eliminado correctamente.")
                st.rerun()

            except Exception as e:
                st.error(f"Error al eliminar paciente: {e}")


if st.session_state.get("mostrar_form_editar_paciente", False):
    with st.expander("✏️ Editar ficha del paciente", expanded=True):
        editar_nombre = st.text_input(
            "Nombre del paciente",
            value=str(paciente_actual.get("nombre", "")),
            key=f"editar_nombre_{paciente_id}"
        )
        sexo_actual = str(paciente_actual.get("sexo", "hombre")).strip().lower()
        indice_sexo = 1 if sexo_actual == "mujer" else 0
        editar_sexo = st.selectbox(
            "Sexo",
            ["hombre", "mujer"],
            index=indice_sexo,
            key=f"editar_sexo_{paciente_id}"
        )

        fecha_nac_actual = pd.to_datetime(
            paciente_actual.get("fecha_nacimiento"),
            errors="coerce"
        )
        if pd.isna(fecha_nac_actual):
            fecha_nac_default = date(1970, 1, 1)
        else:
            fecha_nac_default = fecha_nac_actual.date()

        editar_fecha_nacimiento = st.date_input(
            "Fecha de nacimiento",
            value=fecha_nac_default,
            min_value=date(1920, 1, 1),
            max_value=date.today(),
            format="DD/MM/YYYY",
            key=f"editar_fecha_nacimiento_{paciente_id}"
        )

        talla_actual = pd.to_numeric(paciente_actual.get("talla_m"), errors="coerce")
        if pd.isna(talla_actual) or float(talla_actual) <= 0:
            talla_actual = 1.60

        editar_talla = st.number_input(
            "Talla (m)",
            min_value=0.50,
            max_value=2.50,
            value=round(float(talla_actual), 2),
            step=0.01,
            format="%.2f",
            key=f"editar_talla_{paciente_id}"
        )

        col_ed1, col_ed2 = st.columns(2)

        with col_ed1:
            if st.button("Guardar cambios", key=f"btn_guardar_edicion_{paciente_id}"):
                try:
                    actualizar_paciente(
                        paciente_id=paciente_id,
                        nombre=editar_nombre,
                        sexo=editar_sexo,
                        fecha_nacimiento=editar_fecha_nacimiento,
                        talla_m=editar_talla
                    )
                    st.session_state["mostrar_form_editar_paciente"] = False
                    st.session_state["paciente_nombre_pendiente"] = str(editar_nombre).strip().lower()
                    st.session_state["paciente_id_seleccionado"] = paciente_id
                    st.session_state["paciente_cargado_id"] = None
                    st.success("Ficha del paciente actualizada correctamente.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al actualizar paciente: {e}")

        with col_ed2:
            if st.button("Cancelar edición", key=f"btn_cancelar_edicion_{paciente_id}"):
                st.session_state["mostrar_form_editar_paciente"] = False
                st.rerun()

# =========================================================
# FICHA + DATAFRAMES BASE
# =========================================================
df_peso_export = obtener_historial_peso(paciente_id) if paciente_id is not None else pd.DataFrame()
df_inbody_export = obtener_historial_inbody(paciente_id) if paciente_id is not None else pd.DataFrame()
df_eval_export = obtener_historial_paciente(paciente_id) if paciente_id is not None else pd.DataFrame()
df_medicacion_export = obtener_historial_medicacion(paciente_id) if paciente_id is not None else pd.DataFrame()

if st.session_state.get("paciente_cargado_id") != paciente_id:
    cargar_datos_paciente_en_widgets(paciente_actual, df_peso_export)

ficha = construir_ficha_paciente(paciente_actual, df_eval_export)

ultimo_id_peso = obtener_ultimo_id_peso(df_peso_export)
ultimo_id_inbody = obtener_ultimo_id_inbody(df_inbody_export)

# =========================================================
# EXPORTACIONES ON DEMAND
# =========================================================
excel_state_key = "excel_general_bytes"
pdf_state_key = f"pdf_paciente_{paciente_id}"

with top2:
    if st.button("Preparar Excel", key="btn_preparar_excel"):
        with st.spinner("Generando Excel..."):
            st.session_state[excel_state_key] = generar_excel_general(pacientes).getvalue()

    if excel_state_key in st.session_state:
        st.download_button(
            label="Descargar Excel",
            data=st.session_state[excel_state_key],
            file_name="reporte_pacientes.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="descargar_excel_final"
        )

with top3:
    if st.button("Preparar PDF", key=f"btn_preparar_pdf_{paciente_id}"):
        with st.spinner("Generando PDF..."):
            st.session_state[pdf_state_key] = generar_pdf_paciente(
                ficha,
                df_peso_export,
                df_inbody_export,
                df_eval_export,
                df_medicacion_export
            ).getvalue()

    if pdf_state_key in st.session_state:
        st.download_button(
            label="Descargar PDF",
            data=st.session_state[pdf_state_key],
            file_name=f"{paciente_nombre.replace(' ', '_')}_reporte.pdf",
            mime="application/pdf",
            key=f"descargar_pdf_final_{paciente_id}"
        )

st.divider()

# =========================================================
# TARJETAS RESUMEN
# =========================================================
k1, k2, k3, k4 = st.columns(4)

with k1:
    with st.container(border=True):
        st.markdown("#### Paciente")
        st.write(f"**Nombre:** {ficha['nombre']}")
        st.write(f"**Sexo:** {str(ficha['sexo']).capitalize() if ficha['sexo'] != '-' else '-'}")
        if ficha["talla_m"] is not None:
            st.write(f"**Talla:** {float(ficha['talla_m']):.2f} m")
        else:
            st.write("**Talla:** -")

with k2:
    with st.container(border=True):
        st.markdown("#### Peso / IMC")
        if not df_peso_export.empty:
            df_peso_tmp = df_peso_export.copy()
            df_peso_tmp["fecha"] = pd.to_datetime(df_peso_tmp["fecha"], errors="coerce")
            df_peso_tmp = df_peso_tmp.dropna(subset=["fecha"]).sort_values("fecha", ascending=False)
            ultimo_peso = df_peso_tmp.iloc[0]
            st.write(f"**Peso:** {float(ultimo_peso['peso_kg']):.1f} kg")
            st.write(f"**IMC:** {float(ultimo_peso['imc']):.2f}")
            if pd.notna(ultimo_peso.get('icc')):
                st.write(f"**ICC:** {float(ultimo_peso['icc']):.2f} ({clasificacion_icc(ficha['sexo'], ultimo_peso['icc'])})")
            if pd.notna(ultimo_peso.get('ica')):
                st.write(f"**ICA:** {float(ultimo_peso['ica']):.2f} ({clasificacion_ica(ultimo_peso['ica'])})")
            st.write(f"**Fecha:** {ultimo_peso['fecha'].strftime('%d-%m-%Y')}")
        else:
            st.write("Sin registros")

with k3:
    with st.container(border=True):
        st.markdown("#### Estado corporal")
        if not df_inbody_export.empty and ficha["talla_m"] is not None:
            df_corporal_tmp = enriquecer_historial_corporal(
                df_inbody_export,
                str(ficha["sexo"]).strip().lower(),
                ficha["talla_m"]
            )
            df_corporal_tmp["fecha"] = pd.to_datetime(df_corporal_tmp["fecha"], errors="coerce")
            df_corporal_tmp = df_corporal_tmp.dropna(subset=["fecha"]).sort_values("fecha", ascending=False)
            ultimo_corporal = df_corporal_tmp.iloc[0]

            st.write(f"**Estado:** {ultimo_corporal.get('diagnostico_corporal', '-')}")
            st.write(f"**Grasa:** {ultimo_corporal.get('clasif_grasa', '-')}")
            st.write(f"**Músculo:** {ultimo_corporal.get('clasif_musculo', '-')}")
        else:
            st.write("Sin registros")

with k4:
    with st.container(border=True):
        st.markdown("#### Última evaluación")
        st.write(f"**Fecha:** {ficha['ultima_fecha']}")
        st.write(f"**Prueba:** {ficha['ultima_prueba']}")
        st.write(f"**Clasificación:** {ficha['ultima_clasificacion']}")

st.divider()


# =========================================================
# INGRESO DE DATOS
# =========================================================
st.markdown("## Ingreso de datos")

# =========================================================
# PESO E IMC
# =========================================================
st.markdown("### Peso e IMC")

col_peso_ext_1, col_peso_centro, col_peso_ext_2 = st.columns([1, 4, 1])

with col_peso_centro:
    with st.container(border=True):
        if ficha["talla_m"] is None or float(ficha["talla_m"]) <= 0:
            st.warning("Este paciente no tiene una talla válida cargada en la tabla pacientes.")
        else:
            col_p_izq, col_p_der = st.columns([1.35, 1])

            with col_p_izq:
                fecha_peso = st.date_input(
                    "Fecha de peso",
                    value=date.today(),
                    key=f"fecha_peso_{paciente_id}"
                )

                peso_kg = st.number_input(
                    "Peso (kg)",
                    min_value=0.0,
                    max_value=300.0,
                    step=0.1,
                    format="%.1f",
                    key=f"peso_kg_{paciente_id}"
                )

                cintura_cm = st.number_input(
                    "Cintura (cm)",
                    min_value=0.0,
                    max_value=300.0,
                    step=0.1,
                    format="%.1f",
                    key=f"cintura_cm_{paciente_id}"
                )

                cadera_cm = st.number_input(
                    "Cadera (cm)",
                    min_value=0.0,
                    max_value=300.0,
                    step=0.1,
                    format="%.1f",
                    key=f"cadera_cm_{paciente_id}"
                )

            imc_calculado = round(float(peso_kg) / (float(ficha["talla_m"]) ** 2), 2)
            clasificacion_imc, color_imc = clasificar_imc(imc_calculado)
            icc_calculado = calcular_icc(cintura_cm, cadera_cm)
            ica_calculado = calcular_ica(cintura_cm, ficha["talla_m"])

            with col_p_der:
                st.markdown("#### Resultados")
                st.markdown(f"**IMC:** {imc_calculado:.2f}")
                st.markdown(f"**Clasificación:** {color_imc} {clasificacion_imc}")
                st.markdown(f"**ICC:** {icc_calculado:.2f}" if icc_calculado is not None else "**ICC:** -")
                st.markdown(f"**Riesgo ICC:** {clasificacion_icc(ficha['sexo'], icc_calculado)}")
                st.markdown(f"**ICA:** {ica_calculado:.2f}" if ica_calculado is not None else "**ICA:** -")
                st.markdown(f"**Riesgo ICA:** {clasificacion_ica(ica_calculado)}")

            bp1, bp2 = st.columns(2)

            with bp1:
                if st.button("Guardar peso", key=f"btn_guardar_peso_{paciente_id}"):
                    try:
                        guardar_peso(
                            paciente_id=paciente_id,
                            fecha_medicion=fecha_peso,
                            peso_kg=peso_kg,
                            talla_m=ficha["talla_m"],
                            cintura_cm=cintura_cm,
                            cadera_cm=cadera_cm
                        )
                        st.success("Peso guardado correctamente.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error al guardar peso: {e}")

            with bp2:
                if st.button(
                    "Borrar último peso",
                    key=f"btn_borrar_ultimo_peso_{paciente_id}",
                    disabled=ultimo_id_peso is None
                ):
                    try:
                        eliminar_registro_peso(ultimo_id_peso)
                        st.success("Último registro de peso eliminado.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error al borrar último peso: {e}")



st.markdown("#### Historial de peso / IMC")

df_peso_hist = df_peso_export.copy()

if not df_peso_hist.empty:
    df_peso_hist["fecha"] = pd.to_datetime(df_peso_hist["fecha"], errors="coerce")
    df_peso_hist = df_peso_hist.dropna(subset=["fecha"]).sort_values("fecha", ascending=False)

    st.markdown("**Fecha | Peso | IMC | Cintura | Cadera | ICC | ICA | Eliminar**")

    for _, row in df_peso_hist.iterrows():
        c1, c2, c3, c4, c5, c6, c7, c8 = st.columns([1.2, 0.9, 0.9, 0.9, 0.9, 0.8, 0.8, 0.5])

        fecha_txt = row["fecha"].strftime("%Y-%m-%d") if pd.notna(row.get("fecha")) else ""
        peso_txt = f"{float(row['peso_kg']):.1f}" if pd.notna(row.get("peso_kg")) else ""
        imc_txt = f"{float(row['imc']):.2f}" if pd.notna(row.get("imc")) else ""
        cintura_txt = f"{float(row['cintura_cm']):.1f}" if pd.notna(row.get("cintura_cm")) else ""
        cadera_txt = f"{float(row['cadera_cm']):.1f}" if pd.notna(row.get("cadera_cm")) else ""
        icc_txt = f"{float(row['icc']):.2f}" if pd.notna(row.get("icc")) else ""
        ica_txt = f"{float(row['ica']):.2f}" if pd.notna(row.get("ica")) else ""

        c1.write(fecha_txt)
        c2.write(peso_txt)
        c3.write(imc_txt)
        c4.write(cintura_txt)
        c5.write(cadera_txt)
        c6.write(icc_txt)
        c7.write(ica_txt)

        if c8.button("🗑", key=f"del_peso_{row['id']}"):
            try:
                eliminar_registro_peso(row["id"])
                st.success("Registro de peso eliminado.")
                st.rerun()
            except Exception as e:
                st.error(f"Error al eliminar registro de peso: {e}")
else:
    st.info("Sin historial de peso / IMC.")

st.markdown("#### Evolución de peso e IMC")

df_peso = df_peso_export.copy()

if not df_peso.empty:
    df_peso["fecha"] = pd.to_datetime(df_peso["fecha"], errors="coerce")
    df_peso["peso_kg"] = pd.to_numeric(df_peso["peso_kg"], errors="coerce")
    df_peso["imc"] = pd.to_numeric(df_peso["imc"], errors="coerce")
    df_peso = df_peso.dropna(subset=["fecha", "peso_kg", "imc"]).sort_values("fecha", ascending=True)

    if not df_peso.empty:
        df_peso["fecha_str"] = df_peso["fecha"].dt.strftime("%Y-%m-%d")

        valor_max = max(df_peso["peso_kg"].max(), df_peso["imc"].max())
        valor_max = max(10, round(valor_max * 1.10, 1))

        df_peso_long = df_peso.melt(
            id_vars=["fecha_str"],
            value_vars=["peso_kg", "imc"],
            var_name="variable",
            value_name="valor"
        )

        grafico_doble = (
            alt.Chart(df_peso_long)
            .mark_line(point=True)
            .encode(
                x=alt.X("fecha_str:O", title="Fecha", axis=alt.Axis(labelFontSize=12, titleFontSize=14)),
                y=alt.Y(
                    "valor:Q",
                    title="Valor",
                    scale=alt.Scale(domain=[0, valor_max], nice=False, zero=True),
                    axis=alt.Axis(labelFontSize=12, titleFontSize=14)
                ),
                color=alt.Color("variable:N", title="Serie"),
                tooltip=[
                    alt.Tooltip("fecha_str:O", title="Fecha"),
                    alt.Tooltip("variable:N", title="Serie"),
                    alt.Tooltip("valor:Q", title="Valor", format=".2f")
                ]
            )
            .properties(height=350, width=700)
        )

        col_g1, col_g2, col_g3 = st.columns([1,6,1])
        with col_g2:
            st.altair_chart(grafico_doble, use_container_width=False)
    else:
        st.info("Sin datos válidos de peso / IMC.")
else:
    st.info("Sin datos de peso / IMC.")

st.divider()
st.markdown("### Evaluación funcional")

with st.container(border=True):
    sexo_paciente = str(paciente_actual.get("sexo", "")).strip().lower() if paciente_actual else ""
    sexo = "Mujer" if sexo_paciente == "mujer" else "Hombre"
    edad_real = calcular_edad_desde_fecha(paciente_actual.get("fecha_nacimiento")) if paciente_actual else 0
    talla_m_paciente = paciente_actual.get("talla_m") if paciente_actual else None
    altura_ref_caminata = obtener_altura_referencia_caminata(talla_m_paciente)

    prueba = st.selectbox(
        "Seleccionar prueba",
        ["Caminata 6 minutos", "Prensión manual", "Levantarse de la silla"],
        key="selector_prueba"
    )

    st.write(f"**Sexo del paciente:** {sexo}")
    st.write(f"**Edad del paciente:** {edad_real} años")
    if talla_m_paciente is not None:
        st.write(f"**Talla del paciente:** {float(talla_m_paciente):.2f} m")
    else:
        st.write("**Talla del paciente:** -")

    valor_medido = 0.0
    altura = None

    if prueba == "Caminata 6 minutos":
        altura = altura_ref_caminata
        st.write(f"**Altura de referencia aplicada:** {altura_ref_caminata if altura_ref_caminata is not None else '-'} cm")
        if 40 <= edad_real <= 80:
            edad_ref_preview = min([40, 50, 60, 70, 80], key=lambda x: abs(x - edad_real))
            st.write(f"**Edad de referencia aplicada:** {edad_ref_preview} años")
        else:
            st.write("**Edad de referencia aplicada:** fuera del rango validado")

        valor_medido = st.number_input(
            "Distancia caminada (metros)",
            min_value=0.0,
            max_value=2000.0,
            step=1.0,
            format="%.2f",
            key="valor_caminata"
        )

    elif prueba == "Prensión manual":
        if 20 <= edad_real <= 100:
            st.write(f"**Grupo etario aplicado:** {grupo_edad_prension(edad_real)}")
        else:
            st.write("**Grupo etario aplicado:** fuera del rango validado")

        valor_medido = st.number_input(
            "Fuerza de prensión (kg)",
            min_value=0.0,
            max_value=100.0,
            step=0.1,
            format="%.1f",
            key="valor_prension"
        )

    elif prueba == "Levantarse de la silla":
        grupo_preview = grupo_edad_silla(edad_real)
        st.write(f"**Grupo etario aplicado:** {grupo_preview if grupo_preview is not None else 'No validado (< 65 años)'}")

        valor_medido = st.number_input(
            "Cantidad de repeticiones",
            min_value=0.0,
            max_value=60.0,
            step=1.0,
            format="%.0f",
            key="valor_silla"
        )

    percentil, clasificacion, referencia_p50, referencia_altura, referencia_edad = calcular_resultado(
        prueba=prueba,
        sexo=sexo,
        edad=edad_real,
        altura=altura,
        valor_medido=valor_medido
    )

    if valor_medido > 0:
        color = color_clasificacion(clasificacion)

        st.markdown(
            f"""
            <div style="
                background-color:{color};
                color:white;
                padding:10px 12px;
                border-radius:8px;
                text-align:center;
                font-size:18px;
                font-weight:600;
                margin-top:18px;
                margin-bottom:10px;
            ">
                {clasificacion}
            </div>
            """,
            unsafe_allow_html=True
        )

        st.markdown(
            f"""
            <div style="
                background-color:#dff0e6;
                color:#1b5e20;
                padding:8px 12px;
                border-radius:8px;
                font-size:15px;
                margin-bottom:14px;
            ">
                Percentil estimado: <b>{f'P{percentil}' if percentil is not None else '-'}</b>
            </div>
            """,
            unsafe_allow_html=True
        )

        st.write(f"**Rango percentilar:** {rango_percentilar(percentil)}")
        st.write(f"**Referencia P50:** {referencia_p50}")

        if referencia_altura != "-":
            st.write(f"**Referencia de altura:** {referencia_altura}")

        if referencia_edad != "-":
            st.write(f"**Referencia etaria:** {referencia_edad}")

        st.write(f"**Interpretación clínica:** {interpretacion_clinica(clasificacion)}")
    else:
        st.info("Ingresá el valor logrado para calcular el percentil y la clasificación.")

    if st.button("Guardar evaluación", key="btn_guardar_evaluacion"):
        if not paciente_nombre:
            st.warning("Seleccioná un paciente antes de guardar.")
        elif valor_medido <= 0:
            st.warning("Ingresá un valor logrado mayor a 0.")
        elif percentil is None:
            st.warning(referencia_edad if referencia_edad not in ["-", None, ""] else "No se pudo calcular el percentil con las tablas validadas.")
        else:
            try:
                guardar_evaluacion(
                    paciente_id=paciente_id,
                    paciente_nombre=paciente_nombre,
                    sexo=sexo,
                    edad=edad_real,
                    prueba=prueba,
                    valor_medido=valor_medido,
                    percentil=percentil,
                    clasificacion=clasificacion
                )
                st.success("Evaluación guardada correctamente.")
                st.rerun()
            except Exception as e:
                st.error(f"Error al guardar: {e}")


st.markdown("#### Historial funcional")

df_historial = df_eval_export.copy()

if not df_historial.empty:
    filtro_historial_global = st.selectbox(
        "Filtrar historial por prueba",
        options=["Todas", "Caminata 6 minutos", "Prensión manual", "Levantarse de la silla"],
        index=0,
        key="filtro_historial_prueba"
    )

    prueba_filtro = filtro_historial_global

    if prueba_filtro == "Todas":
        df_historial_filtrado = df_historial.copy()
    else:
        df_historial_filtrado = df_historial[
            df_historial["prueba"].astype(str).str.strip() == prueba_filtro
        ].copy()

    columnas_mostrar = ["id", "fecha", "prueba", "valor_medido", "percentil", "clasificacion"]
    columnas_existentes = [c for c in columnas_mostrar if c in df_historial_filtrado.columns]

    df_historial_mostrar = df_historial_filtrado[columnas_existentes].copy()

    if "fecha" in df_historial_mostrar.columns:
        df_historial_mostrar["fecha"] = pd.to_datetime(
            df_historial_mostrar["fecha"],
            errors="coerce"
        ).dt.strftime("%Y-%m-%d")

    df_historial_mostrar = df_historial_mostrar.sort_values(by="fecha", ascending=False)

    st.markdown("**Fecha | Prueba | Valor | Percentil | Clasificación | Eliminar**")

    for _, row in df_historial_mostrar.iterrows():
        c1, c2, c3, c4, c5, c6 = st.columns([1, 2, 1, 1, 1, 0.5])

        c1.write(row.get("fecha", ""))
        c2.write(row.get("prueba", ""))
        c3.write(row.get("valor_medido", ""))
        c4.write(row.get("percentil", ""))
        c5.write(row.get("clasificacion", ""))

        if c6.button("🗑", key=f"del_{row['id']}"):
            try:
                eliminar_evaluacion(row["id"])
                st.success("Evaluación eliminada.")
                st.rerun()
            except Exception as e:
                st.error(f"Error al eliminar: {e}")

    st.markdown("#### Evolución del percentil funcional")

    df_graf_base = df_eval_export.copy()

    if not df_graf_base.empty and {"fecha", "percentil", "prueba", "clasificacion"}.issubset(df_graf_base.columns):
        df_graf_base["fecha"] = pd.to_datetime(df_graf_base["fecha"], errors="coerce")
        df_graf_base["percentil"] = pd.to_numeric(df_graf_base["percentil"], errors="coerce")
        df_graf_base["prueba"] = df_graf_base["prueba"].astype(str).str.strip()
        df_graf_base["clasificacion"] = df_graf_base["clasificacion"].astype(str).str.strip()
        df_graf_base = df_graf_base.dropna(subset=["fecha", "percentil", "prueba"])

        filtro_historial = st.session_state.get("filtro_historial_prueba", "Todas")
        opciones_prueba = ["Caminata 6 minutos", "Prensión manual", "Levantarse de la silla"]

        if filtro_historial != "Todas":
            prueba_grafico = filtro_historial
            st.markdown(f"**Prueba para gráfico:** {prueba_grafico}")
        else:
            prueba_grafico = st.selectbox(
                "Prueba para gráfico",
                options=opciones_prueba,
                key="selector_grafico_prueba"
            )

        df_prueba = df_graf_base[df_graf_base["prueba"] == prueba_grafico].copy()

        if not df_prueba.empty:
            df_prueba = (
                df_prueba.sort_values("fecha")
                .groupby("fecha", as_index=False)
                .agg({
                    "percentil": "mean",
                    "clasificacion": "last"
                })
            )

            df_prueba["fecha_str"] = df_prueba["fecha"].dt.strftime("%Y-%m-%d")
            df_prueba["Etiqueta"] = df_prueba["percentil"].apply(lambda x: f"P{round(x, 1)}")

            dominio_x = df_prueba["fecha_str"].tolist()

            linea_p50_df = pd.DataFrame({
                "fecha_str": dominio_x,
                "p50": [50] * len(dominio_x)
            })

            linea_p50 = alt.Chart(linea_p50_df).mark_line(strokeDash=[6, 4]).encode(
                x=alt.X("fecha_str:O", title="Fecha", sort=dominio_x),
                y=alt.Y(
                    "p50:Q",
                    title="Percentil",
                    scale=alt.Scale(domain=[0, 100], nice=False, zero=True)
                ),
                tooltip=[alt.Tooltip("p50:Q", title="Referencia", format=".0f")]
            )

            linea = alt.Chart(df_prueba).mark_line().encode(
                x=alt.X("fecha_str:O", title="Fecha", sort=dominio_x),
                y=alt.Y(
                    "percentil:Q",
                    title="Percentil",
                    scale=alt.Scale(domain=[0, 100], nice=False, zero=True)
                ),
                tooltip=[
                    alt.Tooltip("fecha_str:O", title="Fecha"),
                    alt.Tooltip("percentil:Q", title="Percentil", format=".1f"),
                    alt.Tooltip("clasificacion:N", title="Clasificación")
                ]
            )

            puntos = alt.Chart(df_prueba).mark_circle(size=110).encode(
                x=alt.X("fecha_str:O", sort=dominio_x),
                y=alt.Y(
                    "percentil:Q",
                    scale=alt.Scale(domain=[0, 100], nice=False, zero=True)
                ),
                color=alt.Color(
                    "clasificacion:N",
                    title="Clasificación",
                    scale=alt.Scale(
                        domain=["Muy bajo", "Bajo", "Ligeramente bajo", "Normal", "Bueno", "Muy bueno"],
                        range=["#d32f2f", "#f57c00", "#fbc02d", "#388e3c", "#1976d2", "#00796b"]
                    )
                ),
                tooltip=[
                    alt.Tooltip("fecha_str:O", title="Fecha"),
                    alt.Tooltip("percentil:Q", title="Percentil", format=".1f"),
                    alt.Tooltip("clasificacion:N", title="Clasificación")
                ]
            )

            etiquetas = alt.Chart(df_prueba).mark_text(
                dy=-12,
                fontSize=12
            ).encode(
                x=alt.X("fecha_str:O", sort=dominio_x),
                y=alt.Y(
                    "percentil:Q",
                    scale=alt.Scale(domain=[0, 100], nice=False, zero=True)
                ),
                text="Etiqueta:N"
            )

            grafico = (linea_p50 + linea + puntos + etiquetas).properties(height=280)
            st.altair_chart(grafico, use_container_width=True)
        else:
            st.info("Sin datos funcionales para esa prueba.")
    else:
        st.info("Sin historial funcional para graficar.")
else:
    st.info("Sin historial funcional.")

st.divider()
st.markdown("### Composición corporal")

with st.container(border=True):
    if ficha["talla_m"] is None or float(ficha["talla_m"]) <= 0:
        st.warning("Para cargar composición corporal primero hay que tener una talla válida en el paciente.")
    else:
        sexo_corporal = str(ficha["sexo"]).strip().lower()

        col_c1, col_c2 = st.columns(2)

        with col_c1:
            fecha_inbody = st.date_input("Fecha estudio", value=date.today(), key=f"inbody_fecha_{paciente_id}")
            peso_inbody = st.number_input("Peso (kg)", min_value=0.0, max_value=300.0, step=0.1, key=f"inbody_peso_{paciente_id}")
            imc_inbody_calc = round(float(peso_inbody) / (float(ficha["talla_m"]) ** 2), 2) if peso_inbody and ficha["talla_m"] else 0.0
            st.markdown(f"**IMC calculado:** {imc_inbody_calc:.2f}")
            grasa_pct = st.number_input("% grasa corporal", min_value=0.0, max_value=80.0, step=0.1, key=f"inbody_grasa_{paciente_id}")

        with col_c2:
            masa_muscular = st.number_input("Masa muscular (kg)", min_value=0.0, max_value=100.0, step=0.1, key=f"inbody_musculo_{paciente_id}")
            agua_pct = st.number_input("% agua corporal", min_value=0.0, max_value=100.0, step=0.1, key=f"inbody_agua_{paciente_id}")
            grasa_visceral = st.number_input("Grasa visceral", min_value=0.0, max_value=30.0, step=0.1, key=f"inbody_visceral_{paciente_id}")
            metabolismo = st.number_input("Metabolismo basal", min_value=0.0, max_value=4000.0, step=10.0, key=f"inbody_metabolismo_{paciente_id}")

        observaciones_inbody = st.text_area("Observaciones", key=f"inbody_obs_{paciente_id}", height=80)

        resultado_corporal = evaluar_perfil_morfofuncional(
            sexo=sexo_corporal,
            peso_kg=peso_inbody,
            talla_m=ficha["talla_m"],
            grasa_pct=grasa_pct,
            masa_muscular_kg=masa_muscular,
            agua_pct=agua_pct,
            grasa_visceral=grasa_visceral
        )

        bg_estado, fg_estado = color_estado_corporal(resultado_corporal["estado"])

        st.markdown(
            f"""
            <div class="result-card" style="background-color:{bg_estado}; color:{fg_estado};">
                Diagnóstico corporal: {resultado_corporal["estado"]}
            </div>
            """,
            unsafe_allow_html=True
        )

        col_r1, col_r2, col_r3 = st.columns(3)
        with col_r1:
            st.write(f"**IMC:** {resultado_corporal['imc'] if resultado_corporal['imc'] is not None else '-'}")
            st.write(f"**Clasificación IMC:** {resultado_corporal['clasif_imc']}")
            st.write(f"**% grasa:** {resultado_corporal['clasif_grasa']}")

        with col_r2:
            st.write(f"**% agua corporal:** {resultado_corporal['clasif_agua']}")
            st.write(f"**Grasa visceral:** {resultado_corporal['clasif_visceral']}")
            st.write(f"**Músculo relativo %:** {resultado_corporal['musculo_rel_pct'] if resultado_corporal['musculo_rel_pct'] is not None else '-'}")

        with col_r3:
            st.write(f"**Clasificación muscular:** {resultado_corporal['clasif_musculo']}")
            st.write(f"**Metabolismo basal:** {metabolismo if metabolismo is not None else '-'}")
            st.write(f"**Sexo de referencia:** {sexo_corporal.capitalize() if sexo_corporal else '-'}")

        motivos_texto = resultado_corporal["motivos"] if resultado_corporal["motivos"] else ["Sin hallazgos relevantes"]
        motivos_html = "".join([f"<li>{m}</li>" for m in motivos_texto])

        st.markdown(
            f"""
            <div class="motivo-box">
                <b>Motivos:</b>
                <ul style="margin-top:8px; margin-bottom:0;">
                    {motivos_html}
                </ul>
            </div>
            """,
            unsafe_allow_html=True
        )

        st.markdown(
            f"""
            <div class="reco-box">
                <b>Sugerencia:</b><br>
                {resultado_corporal["recomendacion"]}
            </div>
            """,
            unsafe_allow_html=True
        )

        bc1, bc2 = st.columns(2)

        with bc1:
            if st.button("Guardar composición corporal", key=f"guardar_inbody_{paciente_id}"):
                try:
                    guardar_inbody(
                        paciente_id=paciente_id,
                        fecha_estudio=fecha_inbody,
                        peso_kg=peso_inbody,
                        talla_m=ficha["talla_m"],
                        grasa_corporal_pct=grasa_pct,
                        masa_muscular_kg=masa_muscular,
                        agua_corporal_pct=agua_pct,
                        grasa_visceral=grasa_visceral,
                        metabolismo_basal=metabolismo,
                        observaciones=observaciones_inbody
                    )
                    st.success("Composición corporal guardada correctamente")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al guardar composición corporal: {e}")

        with bc2:
            if st.button(
                "Borrar última composición",
                key=f"btn_borrar_ultimo_inbody_{paciente_id}",
                disabled=ultimo_id_inbody is None
            ):
                try:
                    eliminar_registro_corporal(ultimo_id_inbody)
                    st.success("Último registro corporal eliminado.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al borrar última composición: {e}")

# =========================================================
# MEDICACIÓN
# =========================================================


st.markdown("#### Historial corporal")

df_inbody = df_inbody_export.copy()
if not df_inbody.empty:
    df_inbody = enriquecer_historial_corporal(
        df_inbody,
        str(ficha["sexo"]).strip().lower(),
        ficha["talla_m"]
    )
    df_inbody["fecha"] = pd.to_datetime(df_inbody["fecha"], errors="coerce")
    df_inbody = df_inbody.dropna(subset=["fecha"]).sort_values("fecha", ascending=False)

    st.markdown("**Fecha | Peso | IMC | % Grasa | Músculo | Diagnóstico | Eliminar**")

    for _, row in df_inbody.iterrows():
        c1, c2, c3, c4, c5, c6, c7 = st.columns([1.2, 1, 0.8, 1, 1, 1.8, 0.5])

        fecha_txt = row["fecha"].strftime("%Y-%m-%d") if pd.notna(row.get("fecha")) else ""
        peso_txt = f"{float(row['peso_kg']):.1f}" if pd.notna(row.get("peso_kg")) else ""
        imc_txt = f"{float(row['imc']):.2f}" if pd.notna(row.get("imc")) else ""
        grasa_txt = f"{float(row['grasa_corporal_pct']):.1f}" if pd.notna(row.get("grasa_corporal_pct")) else ""
        musculo_txt = f"{float(row['masa_muscular_kg']):.1f}" if pd.notna(row.get("masa_muscular_kg")) else ""
        diagnostico_txt = str(row.get("diagnostico_corporal", ""))

        c1.write(fecha_txt)
        c2.write(peso_txt)
        c3.write(imc_txt)
        c4.write(grasa_txt)
        c5.write(musculo_txt)
        c6.write(diagnostico_txt)

        if c7.button("🗑", key=f"del_inbody_{row['id']}"):
            try:
                eliminar_registro_corporal(row["id"])
                st.success("Registro corporal eliminado.")
                st.rerun()
            except Exception as e:
                st.error(f"Error al eliminar registro corporal: {e}")
else:
    st.info("Sin historial corporal.")

st.divider()
st.markdown("### Medicación")

with st.container(border=True):
    fecha_cambio = st.date_input(
        "Fecha de cambio",
        value=date.today(),
        key=f"med_fecha_{paciente_id}"
    )

    col_m1, col_m2 = st.columns(2)

    with col_m1:
        droga = st.text_input("Droga", key=f"med_droga_{paciente_id}")
        dosis = st.number_input(
            "Dosis",
            min_value=0.0,
            max_value=99999.0,
            step=0.1,
            format="%.2f",
            key=f"med_dosis_{paciente_id}"
        )
        unidad = st.text_input(
            "Unidad",
            placeholder="mg / mcg / ml / comprimidos",
            key=f"med_unidad_{paciente_id}"
        )

    with col_m2:
        frecuencia = st.text_input(
            "Frecuencia",
            placeholder="Cada 24 h / Cada 12 h",
            key=f"med_frecuencia_{paciente_id}"
        )
        via_administracion = st.selectbox(
            "Vía de administración",
            ["Oral", "Subcutánea", "Intravenosa", "Intramuscular", "Tópica", "Inhalatoria", "Otra"],
            key=f"med_via_{paciente_id}"
        )
        estado_medicacion = st.selectbox(
            "Estado",
            ["Activa", "Modificada", "Suspendida"],
            key=f"med_estado_{paciente_id}"
        )

    observaciones_medicacion = st.text_area(
        "Observaciones",
        key=f"med_obs_{paciente_id}",
        height=80
    )

    if st.button("Guardar medicación", key=f"btn_guardar_medicacion_{paciente_id}"):
        try:
            guardar_medicacion(
                paciente_id=paciente_id,
                fecha_cambio=fecha_cambio,
                droga=droga,
                dosis=dosis,
                unidad=unidad,
                frecuencia=frecuencia,
                via_administracion=via_administracion,
                estado=estado_medicacion,
                observaciones=observaciones_medicacion
            )
            st.success("Medicación guardada correctamente.")
            st.rerun()
        except Exception as e:
            st.error(f"Error al guardar medicación: {e}")

    st.markdown("#### Historial de medicación")

    df_medicacion = df_medicacion_export.copy()

    if not df_medicacion.empty:
        if "fecha_cambio" in df_medicacion.columns:
            df_medicacion["fecha_cambio"] = pd.to_datetime(df_medicacion["fecha_cambio"], errors="coerce")
            df_medicacion = df_medicacion.sort_values("fecha_cambio", ascending=False)
            df_medicacion["fecha_cambio"] = df_medicacion["fecha_cambio"].dt.strftime("%Y-%m-%d")

        columnas_medicacion = [
            "fecha_cambio",
            "droga",
            "dosis",
            "unidad",
            "frecuencia",
            "via_administracion",
            "estado",
            "observaciones"
        ]
        columnas_medicacion = [c for c in columnas_medicacion if c in df_medicacion.columns]

        st.dataframe(
            df_medicacion[columnas_medicacion],
            use_container_width=True,
            hide_index=True,
            height=220
        )
    else:
        st.info("Sin historial de medicación.")

st.divider()
filtro_historial_global = "Todas"



st.divider()

# =========================================================
# INFORME INTEGRADO
# =========================================================
informe_integrado = generar_informe_integrado_paciente(
    ficha=ficha,
    df_peso=df_peso_export,
    df_inbody=df_inbody_export,
    df_eval=df_eval_export,
    df_medicacion=df_medicacion_export
)

st.markdown("## Informe integrado del paciente")

bg_estado_global, fg_estado_global = color_estado_global_informe(informe_integrado["estado_global"])

st.markdown(
    f"""
    <div class="result-card" style="background-color:{bg_estado_global}; color:{fg_estado_global};">
        Estado global: {informe_integrado["estado_global"]}
    </div>
    """,
    unsafe_allow_html=True
)

ii1, ii2, ii3 = st.columns(3)

with ii1:
    with st.container(border=True):
        st.markdown("#### Funcional")
        percentiles_funcionales = informe_integrado.get("percentiles_funcionales", {})

        p_caminata = percentiles_funcionales.get("Caminata 6 minutos")
        p_prension = percentiles_funcionales.get("Prensión manual")
        p_silla = percentiles_funcionales.get("Levantarse de la silla")

        st.write(f"**Caminata 6 min:** {f'P{p_caminata}' if p_caminata is not None else '-'}")
        st.write(f"**Prensión manual:** {f'P{p_prension}' if p_prension is not None else '-'}")
        st.write(f"**Levantarse silla:** {f'P{p_silla}' if p_silla is not None else '-'}")

with ii2:
    with st.container(border=True):
        st.markdown("#### Corporal")
        st.write(f"**Estado corporal:** {informe_integrado['estado_corporal']}")
        st.write(f"**Paciente:** {ficha['nombre']}")
        st.write(f"**Sexo / talla:** {str(ficha['sexo']).capitalize() if ficha['sexo'] else '-'} / {ficha['talla_m'] if ficha['talla_m'] is not None else '-'}")

with ii3:
    with st.container(border=True):
        st.markdown("#### Conclusión")
        st.write(f"**Estado integrado:** {informe_integrado['estado_global']}")
        st.write(f"**Última prueba:** {ficha['ultima_prueba']}")
        st.write(f"**Última clasificación:** {ficha['ultima_clasificacion']}")

df_tabla_resumen = informe_integrado["tabla_resumen"]

if df_tabla_resumen is not None and not df_tabla_resumen.empty:
    st.markdown("### Resumen unificado de percentiles y clasificaciones")
    st.dataframe(
        df_tabla_resumen,
        use_container_width=True,
        hide_index=True
    )

st.markdown(
    f"""
    <div class="motivo-box">
        <b>Comentario clínico unificado:</b><br><br>
        {informe_integrado["comentario_unificado"]}
    </div>
    """,
    unsafe_allow_html=True
)

st.markdown(
    f"""
    <div class="reco-box">
        <b>Recomendación unificada:</b><br>
        {informe_integrado["recomendacion_final"]}
    </div>
    """,
    unsafe_allow_html=True
)

st.divider()